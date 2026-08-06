"""
Automatización de Extracción de Datos de Facturas Electrónicas
y Carga en Plantilla de Siigo (Importación de Productos).

═══════════════════════════════════════════════════════════════
ATENCIÓN — DOS FLUJOS COMPLETAMENTE INDEPENDIENTES EN SIIGO:
═══════════════════════════════════════════════════════════════

  FLUJO A — REGISTRO DE COMPRA (este módulo):
    Solo aplica a proveedores de MATERIAS PRIMAS que se inventarían.
    Gmail → ZIP → XML DIAN → codificar → XML McKenna → importar en SIIGO
    como "Crear compra o gasto desde XML o ZIP".

    NO aplica a: consumibles, gastos de envío, servicios generales.
    Esos se registran directamente en SIIGO como gasto/costo sin
    pasar por este proceso de codificación de inventario.

  FLUJO B — FACTURACIÓN DE VENTA (módulo separado: siigo.py):
    Cuando un cliente compra, se crea la factura de venta en SIIGO.
    Ese proceso es completamente distinto y no tiene relación con
    el registro de compra de este módulo.

═══════════════════════════════════════════════════════════════

Pipeline de este módulo (solo proveedores especiales):
  Gmail "FACTURAS MCKG" → ZIP → XML DIAN
    → verificar si proveedor está en lista de proveedores especiales
    → extraer productos + cantidades + IVA
    → generar código McKenna (3init×3palabras + unidad)
    → convertir a unidad mínima (mL / g / Un)
    → calcular precio/unidad con IVA proporcional
    → verificar duplicados en SIIGO
    → generar Excel de registro de productos
    → generar XML de compra con códigos SIIGO internos
    → notificar por WhatsApp con protocolo de carga en SIIGO
"""

import os
import re
import json
import threading
import unicodedata
import requests
import xml.etree.ElementTree as ET
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.tools.sincronizar_facturas_de_compra_siigo import (
    GmailAuthError,
    get_gmail_service,
    leer_correos_no_descargados,
    descargar_y_extraer_zip,
    extraer_datos_xml_dian,
    CARPETA_FACTURAS_LOCAL,
)
from app.services.siigo import autenticar_siigo, PARTNER_ID
from app.utils import enviar_whatsapp_reporte, enviar_whatsapp_archivo

GRUPO_COMPRAS = os.getenv("GRUPO_FACTURACION_COMPRAS_WA", "120363408323873426@g.us")

# ─────────────────────────────────────────────
#  Configuración
# ─────────────────────────────────────────────

CARPETA_IMPORTACIONES = os.path.join("/home/mckg/mi-agente", "importaciones_productos")
os.makedirs(CARPETA_IMPORTACIONES, exist_ok=True)

# Palabras que se ignoran al generar el código del producto
STOPWORDS = {
    'de', 'del', 'la', 'el', 'los', 'las', 'un', 'una', 'para', 'con',
    'por', 'a', 'en', 'al', 'y', 'e', 'o', 'u', 'al', 'se',
}

# Palabras de cantidad/unidad que NO forman parte del código del producto.
# Estas palabras ya determinan el sufijo (g / mL / Un) mediante las funciones
# _extraer_masa_g_descripcion y _extraer_volumen_ml_descripcion.
# Colores (NEGRO, AZUL, ROJO…) y medidas dimensionales (77MM, 30CM…) SÍ se incluyen.
PALABRAS_MEDIDA = {
    # Masa
    'KILO', 'KILOS', 'KILOGRAMO', 'KILOGRAMOS', 'KG', 'KGS',
    'GRAMO', 'GRAMOS', 'GR', 'GRS',
    'MILIGRAMO', 'MILIGRAMOS', 'MG',
    'LIBRA', 'LIBRAS', 'LB', 'LBS',
    'ONZA', 'ONZAS', 'OZ',
    # Volumen
    'LITRO', 'LITROS', 'LT', 'LTS', 'LTR',
    'MILILITRO', 'MILILITROS', 'ML',
    'CENTILITRO', 'CENTILITROS', 'CL', 'CC',
    'GALON', 'GALONES', 'GALO',
    # Unidades genéricas
    'UNIDAD', 'UNIDADES', 'UND', 'UDS', 'PCS', 'PZA', 'PIEZA', 'PIEZAS',
    'BULTO', 'BOLSA', 'CAJA', 'PAQUETE', 'ROLLO',
}

# Envases o material de empaque que se compra por unidad aunque el nombre tenga
# capacidad ("FARMA 5ML AMBAR", "FRASCO 30 ML"). En estos casos "5ML" describe
# el tamaño del envase, no una cantidad de líquido inventariable.
PALABRAS_ENVASE_UNIDAD = {
    'FARMA', 'FRASCO', 'FRASCOS', 'ENVASE', 'ENVASES', 'BOTELLA', 'BOTELLAS',
    'GOTERO', 'GOTEROS', 'PIPETA', 'PIPETAS', 'TARRO', 'TARROS', 'POTE',
    'POTES', 'TAPA', 'TAPAS', 'VALVULA', 'VALVULAS', 'DOSIFICADOR',
    'DOSIFICADORES', 'ATOMIZADOR', 'ATOMIZADORES', 'SPRAY', 'ROLLON',
}

# Conversión desde código DIAN de la factura → (unidad_mínima, factor_multiplicador)
# La unidad mínima es la que se registra en SIIGO
CONVERSION_UNIDADES = {
    # Volumen → mL
    'LTR': ('mL', 1_000),        # Litro
    'MLT': ('mL', 1),             # Mililitro
    'CLT': ('mL', 10),            # Centilitro
    'GLL': ('mL', 3_785.41),      # Galón US
    'OZA': ('mL', 29.5735),       # Onza fluida
    # Masa → g
    'KGM': ('g', 1_000),          # Kilogramo
    'GRM': ('g', 1),              # Gramo
    'MGM': ('g', 0.001),          # Miligramo
    'CGM': ('g', 0.01),           # Centigramo
    'LBR': ('g', 453.592),        # Libra
    'ONZ': ('g', 28.3495),        # Onza masa
    # Unidades
    'NAR': ('Un', 1),             # Número de artículo (unidad estándar DIAN)
    'UN':  ('Un', 1),
    'UNI': ('Un', 1),
    'C62': ('Un', 1),             # Another DIAN unit code
    'BX':  ('Un', 1),             # Caja (se registra por caja)
    'PAR': ('Un', 1),             # Par
    'SET': ('Un', 1),             # Set
    'DZN': ('Un', 12),            # Docena → 12 unidades
    'XBX': ('Un', 1),
}

# Código DIAN oficial para la unidad mínima
DIAN_MIN_CODE = {
    'mL': 'MLT',
    'g':  'GRM',
    'Un': 'NAR',
}

# ─────────────────────────────────────────────
#  Lista de proveedores especiales
#  (requieren codificación de inventario McKenna)
# ─────────────────────────────────────────────

_RUTA_PROVEEDORES = os.path.join(
    os.path.dirname(__file__), '..', 'data', 'proveedores_especiales.json'
)

def cargar_proveedores_especiales() -> dict:
    """
    Carga la lista de proveedores que requieren codificación especial de inventario.
    Retorna dict con estructura:
      {
        "proveedores": [
          {"nit": "900123456", "nombre": "PROVEEDOR S.A.S.", "activo": true, "nota": "..."},
          ...
        ]
      }
    Si el archivo no existe, crea uno vacío con comentario de uso.
    """
    if not os.path.exists(_RUTA_PROVEEDORES):
        plantilla = {
            "_instrucciones": (
                "Agrega aquí los NITs de los proveedores que venden MATERIAS PRIMAS "
                "que se deben inventariar en SIIGO con el proceso de codificación McKenna. "
                "Los proveedores NO listados aquí se ignoran en el flujo de importación de productos. "
                "Sus facturas (consumibles, gastos de envío, servicios) se deben registrar "
                "directamente en SIIGO como compra/gasto normal, sin pasar por este módulo."
            ),
            "proveedores": []
        }
        os.makedirs(os.path.dirname(_RUTA_PROVEEDORES), exist_ok=True)
        with open(_RUTA_PROVEEDORES, 'w', encoding='utf-8') as f:
            json.dump(plantilla, f, indent=2, ensure_ascii=False)
        print(f"📋 [IMPORTACIÓN] Creado archivo de proveedores especiales vacío: {_RUTA_PROVEEDORES}")
    try:
        with open(_RUTA_PROVEEDORES, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"⚠️ [IMPORTACIÓN] Error leyendo proveedores especiales: {e}")
        return {"proveedores": []}


def es_proveedor_especial(nit: str, nombre: str = '') -> bool:
    """
    Verifica si un proveedor debe procesarse con codificación especial de inventario.
    Compara por NIT (exacto) o por nombre (parcial, case-insensitive).
    Si la lista está vacía, acepta TODOS los proveedores (comportamiento por defecto
    hasta que se configure la lista).
    """
    data = cargar_proveedores_especiales()
    proveedores = [p for p in data.get('proveedores', []) if p.get('activo', True)]

    # Lista vacía → proveedor desconocido: siempre preguntar al operador
    if not proveedores:
        return False

    nit_limpio = re.sub(r'\D', '', nit or '')
    nombre_up  = (nombre or '').upper().strip()

    for p in proveedores:
        nit_p = re.sub(r'\D', '', p.get('nit', ''))
        if nit_limpio and nit_p and nit_limpio == nit_p:
            return True
        nombre_p = p.get('nombre', '').upper().strip()
        if nombre_up and nombre_p and (nombre_p in nombre_up or nombre_up in nombre_p):
            return True
    return False


# Colores corporativos para Excel
COLOR_HEADER_BG = "1F4E79"   # Azul oscuro McKenna
COLOR_HEADER_FG = "FFFFFF"   # Blanco
COLOR_ALT_ROW   = "D9E2F3"   # Azul muy claro para filas alternas


# ─────────────────────────────────────────────
#  Lógica de generación de código
# ─────────────────────────────────────────────

def _normalizar(texto: str) -> str:
    """Elimina tildes y convierte a mayúsculas."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto.upper())
        if unicodedata.category(c) != 'Mn'
    )


def _fragmento_palabra_codigo(palabra: str) -> str:
    """Solo letras y dígitos — descarta %, comas y otros símbolos del nombre del proveedor."""
    return re.sub(r'[^A-Za-z0-9]', '', palabra)


def _sanitizar_codigo_siigo(codigo: str) -> str:
    """Normaliza un código para la API SIIGO (sin %, espacios ni símbolos raros)."""
    return re.sub(r'[^A-Za-z0-9._-]', '', (codigo or '').strip())


def generar_codigo_producto(nombre: str, unidad_minima: str,
                            codigos_usados: set = None) -> str:
    """
    Genera código de producto SIIGO:
      → 3 primeras letras de cada una de las N palabras clave (sin stopwords)
      → + sufijo de unidad: mL | g | Un
      → Si el código ya está en `codigos_usados`, agrega más palabras hasta
        encontrar uno único (evita colisiones dentro de la misma factura).

    Ejemplo:
      "ACEITE DE RICINO"  + mL  →  ACERICmL
      "GLICERINA VEGETAL REFINADA" + g → GLIVEGREFg
      "GOTERO PIPETA 77MM NEGRO" + Un, con "GOTPIP77MUn" ya usado
        →  GOTPIP77MNEGUn
    """
    nombre_norm = _normalizar(nombre)
    palabras_raw = re.split(r'[\s\-_/,.()+]+', nombre_norm)
    palabras_clave = []
    for p in palabras_raw:
        p = _fragmento_palabra_codigo(p)
        if not p:
            continue
        es_medida_compacta = re.match(
            r'^x?\d+(?:ml|g|gr|grs|kg|kgs|l|lt|lts|cc|oz|lb|lbs)$',
            p,
            re.IGNORECASE,
        )
        if (
            p.lower() in STOPWORDS
            or p.upper() in PALABRAS_MEDIDA
            or p.isdigit()
            or len(p) < 2
        ):
            continue
        # En productos por unidad, una medida compacta puede ser parte esencial
        # de la referencia del envase: FARMA 5ML AMBAR -> FAR5MLAMBUn.
        if es_medida_compacta and unidad_minima != 'Un':
            continue
        palabras_clave.append(p)

    # Fallback si el nombre no tiene palabras útiles
    if not palabras_clave:
        palabras_clave = [nombre_norm[:9]] if nombre_norm else ['PROD']

    # Probar con min(3, disponibles) palabras y aumentar hasta agotar las disponibles
    # Esto evita el UnboundLocalError cuando el nombre tiene < 3 palabras clave
    n_inicio = min(3, len(palabras_clave))
    for n in range(n_inicio, len(palabras_clave) + 1):
        fragmentos = [p[:3] for p in palabras_clave[:n]]
        codigo = ''.join(fragmentos) + unidad_minima
        codigo = _sanitizar_codigo_siigo(codigo)
        if codigos_usados is None or codigo not in codigos_usados:
            return codigo

    # Si aún hay colisión (nombres prácticamente idénticos), añadir sufijo numérico
    base = ''.join(p[:3] for p in palabras_clave) + unidad_minima
    codigo = base
    i = 2
    while codigos_usados and codigo in codigos_usados:
        codigo = base + str(i)
        i += 1
    return _sanitizar_codigo_siigo(codigo)


# ─────────────────────────────────────────────
#  Conversión de unidades
# ─────────────────────────────────────────────

def _extraer_unit_code_de_xml(xml_content: str, descripcion_item: str) -> str:
    """
    Extrae el unitCode DIAN de la InvoiceLine que corresponde a la descripción dada.
    Soporta dos formatos DIAN:
      - Invoice directa: root tag es Invoice con InvoiceLines hijos
      - AttachedDocument: la Invoice está embebida como texto en un nodo <Description>
    Retorna NAR (unidad) si no se puede determinar.
    """
    import xml.etree.ElementTree as ET

    def _tag(e): return e.tag.split('}')[-1]

    def _buscar_invoice_node(root):
        """Retorna el nodo Invoice ya sea directo o embebido como texto."""
        # Caso 1: Invoice es el root o hijo directo
        for elem in root.iter():
            if _tag(elem) == 'Invoice':
                return elem
        # Caso 2: AttachedDocument — Invoice embebida como texto en Description
        for elem in root.iter():
            if _tag(elem) == 'Description':
                txt = (elem.text or '').strip()
                if '<Invoice' in txt or '<inv:Invoice' in txt:
                    try:
                        return ET.fromstring(txt)
                    except ET.ParseError:
                        # Limpiar prefijos no declarados (misma técnica que extraer_datos_xml_dian)
                        clean = re.sub(r'(</?)[a-zA-Z0-9]+:', r'\1', txt)
                        clean = re.sub(r' [a-zA-Z0-9]+:([a-zA-Z0-9]+)=', r' \1=', clean)
                        try:
                            return ET.fromstring(clean)
                        except Exception:
                            pass
        return None

    unidades = {}
    try:
        root = ET.fromstring(xml_content)
        invoice = _buscar_invoice_node(root)
        nodo = invoice if invoice is not None else root
        for elem in nodo.iter():
            if _tag(elem) == 'InvoiceLine':
                desc = unit_code = ''
                for sub in elem.iter():
                    if _tag(sub) == 'Description' and sub.text and not desc:
                        desc = sub.text.strip()
                    if _tag(sub) == 'InvoicedQuantity':
                        unit_code = sub.get('unitCode', '')
                if desc and unit_code:
                    unidades[_normalizar(desc)] = unit_code.upper()
    except Exception:
        pass

    # Búsqueda normalizada (insensible a acentos, mayúsculas)
    desc_norm = _normalizar(descripcion_item)
    if desc_norm in unidades:
        return unidades[desc_norm]
    for k, v in unidades.items():
        if desc_norm[:25] in k or k[:25] in desc_norm:
            return v

    # Default: NAR (unidad). No inferir por nombre — "5ML" en un frasco es el volumen,
    # no la unidad de venta.
    return 'NAR'


def _extraer_referencia_proveedor_de_xml(xml_content: str, descripcion_item: str) -> str:
    """
    Extrae SellersItemIdentification/ID de la InvoiceLine que corresponde a la descripción.
    Retorna cadena vacía si no hay referencia del proveedor.
    """
    def _tag(e):
        return e.tag.split('}')[-1]

    def _buscar_invoice_node(root):
        for elem in root.iter():
            if _tag(elem) == 'Invoice':
                return elem
        for elem in root.iter():
            if _tag(elem) == 'Description':
                txt = (elem.text or '').strip()
                if '<Invoice' in txt or '<inv:Invoice' in txt:
                    try:
                        return ET.fromstring(txt)
                    except ET.ParseError:
                        clean = re.sub(r'(</?)[a-zA-Z0-9]+:', r'\1', txt)
                        clean = re.sub(r' [a-zA-Z0-9]+:([a-zA-Z0-9]+)=', r' \1=', clean)
                        try:
                            return ET.fromstring(clean)
                        except Exception:
                            pass
        return None

    referencias = {}
    try:
        root = ET.fromstring(xml_content)
        invoice = _buscar_invoice_node(root)
        nodo = invoice if invoice is not None else root
        for elem in nodo.iter():
            if _tag(elem) == 'InvoiceLine':
                desc = referencia = ''
                for sub in elem.iter():
                    if _tag(sub) == 'Description' and sub.text and not desc:
                        desc = sub.text.strip()
                    if _tag(sub) == 'SellersItemIdentification':
                        for s2 in sub.iter():
                            if _tag(s2) == 'ID' and s2.text:
                                referencia = s2.text.strip()
                if desc and referencia:
                    referencias[_normalizar(desc)] = referencia
    except Exception:
        pass

    desc_norm = _normalizar(descripcion_item)
    if desc_norm in referencias:
        return referencias[desc_norm]
    for k, v in referencias.items():
        if desc_norm[:25] in k or k[:25] in desc_norm:
            return v
    return ''


def _codigo_siigo_desde_referencia(referencia: str) -> str | None:
    """Si la referencia del proveedor coincide con un producto SIIGO, retorna ese código."""
    referencia = (referencia or '').strip()
    if not referencia:
        return None
    try:
        codigo = _codigo_manual_valido(referencia)
    except ValueError:
        return None
    return codigo if buscar_producto_en_siigo_por_codigo(codigo) else None


def convertir_a_unidad_minima(cantidad: float, unit_code: str) -> tuple[float, str, str]:
    """
    Convierte cantidad y unidad DIAN a unidad mínima de inventario.
    Retorna (cantidad_min, unidad_min_simbolo, codigo_dian_min).
    """
    unit_code = (unit_code or 'NAR').upper().strip()
    unidad_min, factor = CONVERSION_UNIDADES.get(unit_code, ('Un', 1))
    cantidad_min = round(cantidad * factor, 6)
    codigo_dian_min = DIAN_MIN_CODE.get(unidad_min, 'NAR')
    return cantidad_min, unidad_min, codigo_dian_min


# ─────────────────────────────────────────────
#  Volumen líquido en descripción
# ─────────────────────────────────────────────

def _extraer_volumen_ml_descripcion(descripcion: str) -> float:
    """
    Detecta si la descripción de un producto líquido incluye su volumen por unidad.
    Retorna el volumen en mL por unidad facturada, o 0.0 si no se detecta.

    Ejemplos:
      "FRAGANCIA OCEANFRESH X 500 ML"  → 500.0
      "FRAGANCIA LIMON X 500ML"        → 500.0
      "ESENCIA COCO X 1 LT"            → 1000.0
      "ACEITE ESENCIAL 250CC"          → 250.0
      "FRAGANCIA SANDALO GALON"        → 3785.41
      "SHAMPOO X 1000 ML"              → 1000.0
    """
    d = descripcion.upper()

    # Galón → 3785.41 mL
    if re.search(r'\bGALO[NÑ]\b', d):
        return 3785.41

    # NNN mL / NNN CC (centímetros cúbicos = mL)
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*(?:ML|CC)\b', d)
    if m:
        return float(m.group(1).replace(',', '.'))

    # NNN LT / NNN LITRO(S) / NNN L (como palabra sola)
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*(?:LITROS?|LTS?)\b', d)
    if m:
        return round(float(m.group(1).replace(',', '.')) * 1000, 6)

    return 0.0


def _descripcion_es_envase_unitario(descripcion: str) -> bool:
    """True si el texto parece un envase comprado por unidad con capacidad nominal."""
    palabras = set(re.split(r'[\s\-_/,.()+]+', _normalizar(descripcion)))
    return bool(palabras & PALABRAS_ENVASE_UNIDAD)


def _extraer_masa_g_descripcion(descripcion: str) -> float:
    """
    Detecta si la descripción de un producto sólido/polvo indica su masa por unidad facturada.
    Retorna la masa en gramos por unidad, o 0.0 si no se detecta.

    Ejemplos:
      "POTASA CAUSTICA * KILO"       → 1000.0
      "ACIDO CITRICO X 1 KG"         → 1000.0
      "BICARBONATO X 500 G"          → 500.0
      "SULFATO DE ZINC 50 GR"        → 50.0
      "BORAX 25 KG BULTO"            → 25000.0
      "UREA COSMETICA X 1 LIBRA"     → 453.592
    """
    d = descripcion.upper()

    # NNN KG / NNN KILO(S) / NNN KILOGRAMO(S)
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*(?:KILOGRAMOS?|KILOS?|KGS?|KG)\b', d)
    if m:
        return round(float(m.group(1).replace(',', '.')) * 1000, 6)

    # Palabra "KILO" sola (sin número previo) → se asume 1 kg = 1000 g
    if re.search(r'(?<!\d)\bKILO\b', d):
        return 1000.0

    # NNN G / NNN GR / NNN GRS / NNN GRAMO(S)
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*(?:GRAMOS?|GRS?|GR|G)\b', d)
    if m:
        return float(m.group(1).replace(',', '.'))

    # LIBRA(S) con o sin número → 453.592 g por libra
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*LIBRAS?', d)
    if m:
        return round(float(m.group(1).replace(',', '.')) * 453.592, 6)
    if re.search(r'\bLIBRA\b', d):
        return 453.592

    return 0.0


# ─────────────────────────────────────────────
#  Multiplicador de contenido en descripción
# ─────────────────────────────────────────────

def _extraer_multiplicador_descripcion(descripcion: str) -> int:
    """
    Detecta si la descripción de un ítem indica que cada unidad facturada
    contiene N unidades mínimas. Ejemplos:
      "bolsa 100 unidades"         → 100
      "caja x 50 und"              → 50
      "portaguia 12x20 - 100 uds"  → 100   (ignora la dimensión 12x20)
      "set 12 piezas"              → 12
    Retorna 1 si no se detecta ningún multiplicador.

    Orden de búsqueda (de más a menos específico):
      1. «X NNN unidades» / «x NNN und» / «* NNN pcs» (con operador × delante)
      2. «NNN unidades» / «NNN und» / «NNN uds» (número seguido de palabras de unidad)
    Se ignoran dimensiones tipo "12x20" porque no van seguidas de palabra de unidad.
    """
    patrones = [
        # Con operador multiplicador explícito antes del número
        r'[xX\*\/]\s*(\d{2,})\s*(?:unidades?|und\.?|uds?\.?|pcs?|piezas?|units?)\b',
        # Número seguido directo de palabra de unidad (sin dimensión previa)
        r'(?<![0-9xX])\b(\d{2,})\s*(?:unidades?|und\.?|uds?\.?|pcs?|piezas?|units?)\b',
    ]
    for patron in patrones:
        m = re.search(patron, descripcion, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            if n > 1:
                return n
    return 1


# ─────────────────────────────────────────────
#  Cálculo de precio por unidad mínima
# ─────────────────────────────────────────────

def calcular_precio_unitario_min(subtotal_linea: float, iva_linea: float, cantidad_min: float) -> float:
    """
    Precio de venta por unidad mínima = (subtotal + IVA proporcional) / cantidad_min.
    Redondea al entero más cercano (COP no usa decimales en SIIGO).
    """
    if cantidad_min <= 0:
        return 0.0
    total_con_iva = subtotal_linea + iva_linea
    return round(total_con_iva / cantidad_min, 2)


# ─────────────────────────────────────────────
#  Verificación de duplicados en SIIGO
# ─────────────────────────────────────────────

def buscar_producto_en_siigo_por_codigo(codigo: str) -> dict | None:
    """
    Consulta SIIGO API para saber si ya existe un producto con ese código.
    Retorna el producto SIIGO cuando hay coincidencia exacta de código.
    Intenta hasta 2 veces con timeout de 15 s antes de rendirse.
    """
    token = autenticar_siigo()
    if not token:
        print(f"  ⚠️ [SIIGO] Sin token — {codigo} se tratará como producto nuevo")
        return None

    headers = {"Authorization": f"Bearer {token}", "Partner-Id": PARTNER_ID}

    for intento in range(1, 3):
        try:
            res = requests.get(
                f"https://api.siigo.com/v1/products?code={codigo}&page_size=1",
                headers=headers,
                timeout=15,
            )
            if res.status_code == 200:
                results = res.json().get('results', [])
                for p in results:
                    if p.get('code', '').upper() == codigo.upper():
                        return p
                return None
            # Cualquier otro status HTTP: no es duplicado
            return None
        except requests.exceptions.Timeout:
            if intento < 2:
                print(f"  ⏳ [SIIGO] Timeout verificando {codigo}, reintentando...")
            else:
                print(f"  ⚠️ [SIIGO] Timeout al verificar {codigo} — se asume nuevo (revisa duplicados en Excel)")
        except Exception as e:
            print(f"  ⚠️ [SIIGO] Error verificando {codigo}: {e} — se asume nuevo")
            break

    return None


def verificar_producto_en_siigo(codigo: str) -> bool:
    """Retorna True si el código ya existe exactamente en SIIGO."""
    return buscar_producto_en_siigo_por_codigo(codigo) is not None


def _resumen_producto_siigo(producto: dict | None) -> dict | None:
    if not producto:
        return None
    unidad_raw = producto.get('unit', {})
    unidad = unidad_raw.get('name', '') if isinstance(unidad_raw, dict) else str(unidad_raw or '')
    return {
        'codigo': producto.get('code', ''),
        'nombre': producto.get('name', ''),
        'unidad': unidad,
        'activo': producto.get('active', True),
    }


# ─────────────────────────────────────────────
#  Generación del Excel de importación
# ─────────────────────────────────────────────

HEADERS_SIIGO = [
    "Tipo de producto",          # A → P-Producto
    "Categoría de inventarios",  # B → 1
    "Código",                    # C → generado
    "Nombre",                    # D → descripción factura
    "Inventariable",             # E → SI
    "Unidad de medida (Código DIAN)",  # F → MLT | GRM | NAR
    "Precio de venta",           # G → calculado
]


def generar_excel_importacion(productos: list, numero_factura: str) -> str | None:
    """
    Genera el archivo Excel para importación de productos nuevos en SIIGO.
    Retorna la ruta del archivo generado, o None si no hay productos nuevos.
    """
    productos_nuevos = [p for p in productos if not p.get('existe_en_siigo', p.get('duplicado'))]
    if not productos_nuevos:
        return None
    nombre_archivo = f"{numero_factura} registro productos.xlsx"
    ruta = os.path.join(CARPETA_IMPORTACIONES, nombre_archivo)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # — Encabezados con estilo
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
    for col_idx, header in enumerate(HEADERS_SIIGO, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 30

    # — Datos
    alt_fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    for row_idx, p in enumerate(productos_nuevos, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        valores = [
            "P-Producto",
            1,
            p['codigo'],
            p['nombre'],
            "SI",
            p['codigo_dian_min'],
            p['precio_unitario'],
        ]
        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical='center')

    # — Anchos de columna
    anchos = [18, 24, 16, 50, 14, 26, 18]
    for col_idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = ancho

    # — Hoja de instrucciones
    ws_instr = wb.create_sheet("Instrucciones")
    ws_instr['A1'] = "INSTRUCCIONES PARA CARGAR EN SIIGO"
    ws_instr['A1'].font = Font(bold=True, size=13)
    instrucciones = [
        "",
        "1. Ingresa a SIIGO Nube.",
        "2. Ve al módulo de Inventario → Productos.",
        "3. Haz clic en el botón ▶ Importación.",
        "4. En el Paso 2, selecciona este archivo Excel.",
        "5. Verifica la vista previa y confirma la importación.",
        "",
        "NOTA: Solo incluye productos NUEVOS. Los que ya existen en SIIGO",
        "      se registran únicamente en el XML de compra (suman inventario).",
    ]
    for i, linea in enumerate(instrucciones, 2):
        ws_instr[f'A{i}'] = linea
    ws_instr.column_dimensions['A'].width = 65

    wb.save(ruta)
    print(f"📊 [IMPORTACIÓN] Excel generado: {ruta}")
    return ruta


# ─────────────────────────────────────────────
#  Generación del XML de compra para SIIGO
# ─────────────────────────────────────────────

def generar_xml_compra_siigo(datos: dict, productos: list, numero_factura: str) -> str:
    """
    Genera el archivo XML de compra para importar en SIIGO mediante
    "Crear compra o gasto desde un XML o ZIP".

    El XML reemplaza los códigos del proveedor por los códigos internos
    McKenna-SIIGO y las cantidades por los valores en unidad mínima.

    Nomenclatura obligatoria: [Número de Factura] codigos siigo.xml

    Retorna la ruta del archivo generado.
    """
    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
    fecha_factura = datos.get('fecha', fecha_hoy)
    nit_proveedor  = datos.get('nit', '') or datos.get('nit_proveedor', '')
    nombre_prov    = datos.get('proveedor', '')
    moneda         = datos.get('moneda', 'COP')

    # ── Construir árbol XML ──────────────────────────────────────
    root = ET.Element('RegistroCompra')
    root.set('xmlns:mckg', 'https://mckennagroup.co/siigo-import/v1')
    root.set('generadoPor', 'Agente Hugo Garcia - McKenna Group S.A.S.')
    root.set('fecha_generacion', fecha_hoy)

    # Aviso separación de flujos
    aviso = ET.SubElement(root, 'Aviso')
    aviso.text = (
        'FLUJO DE REGISTRO DE COMPRA — Independiente del flujo de facturación de venta. '
        'Este XML es para registrar la compra del proveedor en SIIGO con códigos internos McKenna. '
        'NO confundir con las facturas de venta a clientes (módulo siigo.py).'
    )

    # Cabecera
    cab = ET.SubElement(root, 'Cabecera')
    ET.SubElement(cab, 'NumeroFacturaProveedor').text = numero_factura
    ET.SubElement(cab, 'FechaFactura').text           = str(fecha_factura)
    ET.SubElement(cab, 'Moneda').text                 = moneda

    prov = ET.SubElement(cab, 'Proveedor')
    ET.SubElement(prov, 'NIT').text    = nit_proveedor
    ET.SubElement(prov, 'Nombre').text = nombre_prov

    # Detalle de ítems con códigos McKenna
    detalle = ET.SubElement(root, 'Detalle')
    detalle.set('totalItems', str(len(productos)))

    subtotal_global = 0.0
    iva_global      = 0.0

    for idx, p in enumerate(productos, 1):
        item = ET.SubElement(detalle, 'Item')
        item.set('numero', str(idx))

        ET.SubElement(item, 'CodigoSiigo').text = p['codigo']

        desc_orig = ET.SubElement(item, 'DescripcionOriginalProveedor')
        desc_orig.text = p['nombre']

        unidad = ET.SubElement(item, 'Unidad')
        unidad_orig = ET.SubElement(unidad, 'CantidadOriginal')
        unidad_orig.text = str(p['cantidad_original'])
        unidad_orig.set('codigoDIAN', p['unidad_original'])

        unidad_conv = ET.SubElement(unidad, 'CantidadConvertida')
        unidad_conv.text = str(round(p['cantidad_min'], 6))
        unidad_conv.set('codigoDIAN', p['codigo_dian_min'])
        unidad_conv.set('simbolo',    p['unidad_min'])

        conv_info = ET.SubElement(unidad, 'ReglaConversion')
        multiplicador = p.get('multiplicador', 1)
        factor_base   = CONVERSION_UNIDADES.get(p['unidad_original'].upper(), ('?', 1))[1]
        if multiplicador > 1:
            # Producto con contenido múltiple: ej. "caja de 100 unidades"
            conv_info.text = (
                f"1 {p['unidad_original']} contiene {multiplicador} {p['unidad_min']} "
                f"| {p['cantidad_original']} × {multiplicador} = {round(p['cantidad_min'], 6)} {p['unidad_min']}"
            )
        else:
            conv_info.text = (
                f"1 {p['unidad_original']} = {factor_base} {p['unidad_min']} "
                f"| {p['cantidad_original']} × {factor_base} = {round(p['cantidad_min'], 6)} {p['unidad_min']}"
            )

        precios = ET.SubElement(item, 'Precios')
        precios.set('moneda', moneda)
        ET.SubElement(precios, 'Subtotal').text         = f"{p['subtotal']:.2f}"
        ET.SubElement(precios, 'IVA').text              = f"{p['iva']:.2f}"
        total_item = p['subtotal'] + p['iva']
        ET.SubElement(precios, 'Total').text            = f"{total_item:.2f}"
        ET.SubElement(precios, 'PrecioUnitarioMin').text = f"{p['precio_unitario']:.4f}"
        ET.SubElement(precios, 'PrecioNetoMin').text = f"{p.get('precio_neto', 0):.6f}"
        pu_desc = ET.SubElement(precios, 'PrecioUnitarioMinDesc')
        pu_desc.text = (
            f"Precio por 1 {p['unidad_min']} "
            f"(total {total_item:.2f} COP / {round(p['cantidad_min'], 2)} {p['unidad_min']})"
        )

        if p.get('existe_en_siigo', p.get('duplicado')):
            ET.SubElement(item, 'EstadoSiigo').text = 'EXISTENTE - Compra suma inventario del producto en SIIGO'
        else:
            ET.SubElement(item, 'EstadoSiigo').text = 'NUEVO - Registrar primero en Excel de productos'

        subtotal_global += p['subtotal']
        iva_global      += p['iva']

    # Totales globales
    totales = ET.SubElement(root, 'Totales')
    totales.set('moneda', moneda)
    ET.SubElement(totales, 'Subtotal').text   = f"{subtotal_global:.2f}"
    ET.SubElement(totales, 'TotalIVA').text   = f"{iva_global:.2f}"
    ET.SubElement(totales, 'TotalGeneral').text = f"{subtotal_global + iva_global:.2f}"
    ET.SubElement(totales, 'Verificacion').text = (
        'IMPORTANTE: Confirmar que el TotalGeneral coincide con el total '
        'de la factura física del proveedor antes de asentar en SIIGO.'
    )

    # Protocolo de carga en SIIGO
    protocolo = ET.SubElement(root, 'ProtocoloRegistroSIIGO')
    pasos = [
        ('Paso1', 'Ingresar a SIIGO Nube → módulo Compras o Contabilidad.'),
        ('Paso2', 'Hacer clic en el botón: "Crear compra o gasto desde un XML o ZIP".'),
        ('Paso3', f'Cargar el archivo: {numero_factura} codigos siigo.xml'),
        ('Paso4', 'Verificar que el TotalGeneral en SIIGO coincide con el total de la factura original del proveedor.'),
        ('Paso5', 'Si los valores son correctos, asentar el documento para registrar la compra.'),
        ('Paso6', (
            'NOTA: Este proceso registra la COMPRA de inventario. '
            'Es independiente del proceso de facturación de VENTA a clientes. '
            'No mezclar estos dos flujos.'
        )),
    ]
    for tag, texto in pasos:
        ET.SubElement(protocolo, tag).text = texto

    # ── Serializar con indentación ──────────────────────────────
    _indentar_xml(root)

    nombre_archivo = f"{numero_factura} codigos siigo.xml"
    ruta = os.path.join(CARPETA_IMPORTACIONES, nombre_archivo)

    tree = ET.ElementTree(root)
    with open(ruta, 'wb') as f:
        tree.write(f, encoding='utf-8', xml_declaration=True)

    print(f"📄 [IMPORTACIÓN] XML compra SIIGO generado: {ruta}")
    return ruta


def _indentar_xml(elem, nivel=0):
    """Agrega sangría al árbol XML para que sea legible."""
    sangria = '\n' + '  ' * nivel
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = sangria + '  '
        if not elem.tail or not elem.tail.strip():
            elem.tail = sangria
        for hijo in elem:
            _indentar_xml(hijo, nivel + 1)
        if not hijo.tail or not hijo.tail.strip():
            hijo.tail = sangria
    else:
        if nivel and (not elem.tail or not elem.tail.strip()):
            elem.tail = sangria
    if not nivel:
        elem.tail = '\n'


# ─────────────────────────────────────────────
#  Cola de aprobación (facturas pendientes)
# ─────────────────────────────────────────────

import base64

_RUTA_PENDIENTES = os.path.join(
    os.path.dirname(__file__), '..', 'data', 'facturas_compra_pendientes.json'
)
_RUTA_HISTORIAL = os.path.join(
    os.path.dirname(__file__), '..', 'data', 'facturas_compra_historial.json'
)


def _sufijo_factura(numero_factura: str) -> str:
    """Últimos 4 caracteres alfanuméricos del número de factura (para comando corto)."""
    alnum = re.sub(r'\W', '', numero_factura)
    return alnum[-4:].upper() if len(alnum) >= 4 else alnum.upper()


def _cargar_pendientes() -> dict:
    try:
        if os.path.exists(_RUTA_PENDIENTES):
            with open(_RUTA_PENDIENTES, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {"pendientes": {}}


def _guardar_pendientes(data: dict):
    os.makedirs(os.path.dirname(_RUTA_PENDIENTES), exist_ok=True)
    with open(_RUTA_PENDIENTES, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _encolar_factura(numero_factura: str, datos: dict, xml_content: str,
                     es_nuevo_proveedor: bool) -> str:
    """
    Guarda la factura en la cola de pendientes para aprobación por WhatsApp.
    Retorna el sufijo (código corto) asignado.
    """
    sufijo = _sufijo_factura(numero_factura)
    total = sum(
        item.get('subtotal', 0) + sum(
            imp.get('valor', 0) for imp in item.get('impuestos', [])
        )
        for item in datos.get('items', [])
    )
    state = _cargar_pendientes()
    state['pendientes'][sufijo] = {
        'numero_factura':     numero_factura,
        'proveedor':          datos.get('proveedor', ''),
        'nit':                datos.get('nit_proveedor', ''),
        'es_nuevo_proveedor': es_nuevo_proveedor,
        'items_count':        len(datos.get('items', [])),
        'total':              round(total, 2),
        'fecha':              (datos.get('fecha') or '')[:10],
        'estado':             'esperando_clasificacion' if es_nuevo_proveedor else 'esperando_confirmacion',
        'xml_b64':            base64.b64encode(xml_content.encode('utf-8')).decode('ascii'),
        'datos_json':         json.dumps(datos, ensure_ascii=False, default=str),
        'timestamp':          datetime.now().isoformat(),
    }
    _guardar_pendientes(state)
    return sufijo


def _fecha_pendiente(entrada: dict) -> str:
    """Fecha de factura YYYY-MM-DD desde campos guardados o XML."""
    for key in ("fecha", "fecha_factura"):
        raw = str(entrada.get(key) or "").strip()
        if len(raw) >= 10 and raw[0:4].isdigit():
            return raw[:10]
    try:
        datos = json.loads(entrada.get("datos_json") or "{}")
        raw = str(datos.get("fecha") or "").strip()
        if len(raw) >= 10 and raw[0:4].isdigit():
            return raw[:10]
    except Exception:
        pass
    xml_b64 = entrada.get("xml_b64") or ""
    if xml_b64:
        try:
            text = base64.b64decode(xml_b64).decode("utf-8", errors="ignore")
            m = re.search(r"IssueDate[^>]*>\s*(\d{4}-\d{2}-\d{2})", text)
            if m:
                return m.group(1)
        except Exception:
            pass
    return ""


def listar_pendientes_panel(anio: int | None = None) -> dict:
    """Pendientes para el panel; por defecto solo año en curso."""
    anio_filtro = int(anio) if anio is not None else int(datetime.now().year)
    state = _cargar_pendientes()
    items = []
    for sufijo, e in (state.get("pendientes") or {}).items():
        fecha = _fecha_pendiente(e)
        if fecha and not fecha.startswith(str(anio_filtro)):
            continue
        if not fecha:
            # Sin fecha confiable: no mostrar en cola del año (evita arrastre viejo)
            continue
        items.append({
            "sufijo": sufijo,
            "numero_factura": e.get("numero_factura", ""),
            "proveedor": e.get("proveedor", ""),
            "nit": e.get("nit", ""),
            "es_nuevo_proveedor": e.get("es_nuevo_proveedor", False),
            "items_count": e.get("items_count", 0),
            "total": e.get("total", 0),
            "fecha": fecha,
            "estado": e.get("estado", ""),
        })
    items.sort(key=lambda r: r.get("fecha") or "", reverse=True)
    return {
        "pendientes": items,
        "total": len(items),
        "anio": anio_filtro,
    }


def podar_pendientes_fuera_de_anio(anio: int | None = None) -> dict:
    """Quita de la cola las facturas cuya fecha no es del año indicado."""
    anio_filtro = int(anio) if anio is not None else int(datetime.now().year)
    state = _cargar_pendientes()
    pendientes = state.get("pendientes") or {}
    keep = {}
    quitadas = []
    for sufijo, e in pendientes.items():
        fecha = _fecha_pendiente(e)
        if fecha.startswith(str(anio_filtro)):
            if not e.get("fecha"):
                e = dict(e)
                e["fecha"] = fecha
            keep[sufijo] = e
        else:
            quitadas.append({
                "sufijo": sufijo,
                "numero_factura": e.get("numero_factura"),
                "fecha": fecha or None,
            })
    state["pendientes"] = keep
    _guardar_pendientes(state)
    return {
        "anio": anio_filtro,
        "quedan": len(keep),
        "quitadas": len(quitadas),
        "detalle_quitadas": quitadas[:50],
    }


def _buscar_pendiente(sufijo: str):
    """Retorna (key, entrada) o (None, None) si no hay coincidencia."""
    state = _cargar_pendientes()
    sufijo_up = sufijo.strip().upper()
    if sufijo_up in state['pendientes']:
        return sufijo_up, state['pendientes'][sufijo_up]
    # Búsqueda parcial: el sufijo que el usuario escribió coincide con el final de la clave
    for k, v in state['pendientes'].items():
        if k.endswith(sufijo_up) or sufijo_up.endswith(k):
            return k, v
    return None, None


def _quitar_pendiente(sufijo: str):
    state = _cargar_pendientes()
    state['pendientes'].pop(sufijo, None)
    _guardar_pendientes(state)


def _cargar_historial() -> dict:
    try:
        if os.path.exists(_RUTA_HISTORIAL):
            with open(_RUTA_HISTORIAL, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {'historial': []}


def _guardar_historial(data: dict):
    os.makedirs(os.path.dirname(_RUTA_HISTORIAL), exist_ok=True)
    with open(_RUTA_HISTORIAL, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _float_safe(valor, default=None):
    try:
        if valor is None:
            return default
        if isinstance(valor, str):
            valor = valor.replace(',', '').strip()
        return float(valor)
    except (TypeError, ValueError):
        return default


def _precio_referencia_item(item: dict) -> float | None:
    """Precio unitario mínimo de compra (neto sin IVA) para comparar tendencias."""
    for campo in ('precio_neto', 'precio_unitario'):
        val = _float_safe(item.get(campo))
        if val is not None and val > 0:
            return val
    return None


def _tendencia_precio(actual: float, anterior: float) -> str:
    if anterior <= 0:
        return 'nuevo'
    diff_pct = abs(actual - anterior) / anterior
    if diff_pct < 0.005:
        return 'igual'
    return 'subio' if actual > anterior else 'bajo'


def _resumen_items_historial(items: list | None) -> list:
    resumen = []
    for p in items or []:
        if not isinstance(p, dict):
            continue
        precio_neto = _float_safe(p.get('precio_neto'))
        precio_unitario = _float_safe(p.get('precio_unitario'))
        precio_proveedor = _float_safe(p.get('precio_proveedor'))
        resumen.append({
            'nombre': str(p.get('nombre', ''))[:120],
            'codigo': str(p.get('codigo', '')),
            'cantidad_min': p.get('cantidad_min'),
            'unidad_min': p.get('unidad_min', ''),
            'precio_neto': precio_neto,
            'precio_unitario': precio_unitario,
            'precio_proveedor': precio_proveedor,
            'existe_en_siigo': bool(p.get('existe_en_siigo', p.get('duplicado'))),
        })
    return resumen


def registrar_factura_historial(
    entrada: dict,
    accion: str,
    *,
    sufijo: str = '',
    origen: str = 'panel',
    estado: str = 'ok',
    datos: dict | None = None,
    items_resumen: list | None = None,
    nuevos: int = 0,
    en_siigo: int = 0,
    ruta_excel: str | None = None,
    ruta_xml: str | None = None,
    siigo_id: str | None = None,
    mensaje: str = '',
) -> dict:
    """Guarda una factura procesada en el historial local del panel."""
    datos = datos or json.loads(entrada.get('datos_json') or '{}')
    sufijo_limpio = (sufijo or '').strip().upper()
    numero = str(entrada.get('numero_factura') or datos.get('number') or '')
    registro = {
        'id': f"{datetime.now().strftime('%Y%m%d%H%M%S%f')[:17]}-{sufijo_limpio or _sufijo_factura(numero)}",
        'sufijo': sufijo_limpio,
        'numero_factura': numero,
        'proveedor': entrada.get('proveedor') or datos.get('proveedor', ''),
        'nit': entrada.get('nit') or datos.get('nit_proveedor') or datos.get('nit', ''),
        'total': entrada.get('total') or datos.get('total_neto') or 0,
        'fecha_factura': datos.get('fecha', ''),
        'items_count': entrada.get('items_count') or len(datos.get('items') or []),
        'accion': accion,
        'estado': estado,
        'origen': origen,
        'nuevos': int(nuevos or 0),
        'en_siigo': int(en_siigo or 0),
        'items_resumen': items_resumen or [],
        'ruta_excel': os.path.basename(ruta_excel) if ruta_excel else None,
        'ruta_xml': os.path.basename(ruta_xml) if ruta_xml else None,
        'siigo_id': siigo_id,
        'mensaje': (mensaje or '')[:500],
        'timestamp': datetime.now().isoformat(),
    }
    state = _cargar_historial()
    historial = state.setdefault('historial', [])
    historial.insert(0, registro)
    state['historial'] = historial[:500]
    _guardar_historial(state)
    return registro


def _parsear_xml_historial_importacion(ruta_xml: str) -> dict | None:
    """Extrae metadatos de un XML McKenna en importaciones_productos/."""
    try:
        root = ET.parse(ruta_xml).getroot()
        cab = root.find("Cabecera")
        if cab is None:
            return None
        numero = (cab.findtext("NumeroFacturaProveedor") or "").strip()
        if not numero:
            return None
        proveedor_el = cab.find("Proveedor") or cab
        totales = root.find("Totales")
        items_resumen = []
        detalle = root.find("Detalle")
        nuevos = 0
        en_siigo = 0
        if detalle is not None:
            for item in detalle.findall("Item"):
                estado = (item.findtext("EstadoSiigo") or "").strip().upper()
                es_nuevo = estado == "NUEVO"
                if es_nuevo:
                    nuevos += 1
                else:
                    en_siigo += 1
                unidad_el = item.find("Unidad")
                cant_min_el = unidad_el.find("CantidadConvertida") if unidad_el is not None else None
                cantidad_min = _float_safe(cant_min_el.text if cant_min_el is not None else None)
                unidad_min = (cant_min_el.get("simbolo") or "") if cant_min_el is not None else ""
                precios_el = item.find("Precios")
                subtotal = _float_safe(precios_el.findtext("Subtotal") if precios_el is not None else None, 0)
                iva = _float_safe(precios_el.findtext("IVA") if precios_el is not None else None, 0)
                precio_unitario = _float_safe(
                    precios_el.findtext("PrecioUnitarioMin") if precios_el is not None else None
                )
                precio_neto = _float_safe(
                    precios_el.findtext("PrecioNetoMin") if precios_el is not None else None
                )
                if precio_neto is None and cantidad_min and cantidad_min > 0:
                    precio_neto = round(subtotal / cantidad_min, 6)
                items_resumen.append({
                    "nombre": str(item.findtext("DescripcionOriginalProveedor") or "")[:120],
                    "codigo": str(item.findtext("CodigoSiigo") or ""),
                    "cantidad_min": cantidad_min,
                    "unidad_min": unidad_min,
                    "precio_neto": precio_neto,
                    "precio_unitario": precio_unitario,
                    "precio_proveedor": None,
                    "existe_en_siigo": not es_nuevo,
                })
        total_txt = totales.findtext("TotalGeneral") if totales is not None else None
        try:
            total = float(str(total_txt).replace(",", "")) if total_txt else 0.0
        except (TypeError, ValueError):
            total = 0.0
        return {
            "numero_factura": numero,
            "fecha_factura": (cab.findtext("FechaFactura") or "")[:10],
            "proveedor": (proveedor_el.findtext("Nombre") or "").strip(),
            "nit": (proveedor_el.findtext("NIT") or "").strip(),
            "total": total,
            "items_count": len(items_resumen),
            "items_resumen": items_resumen,
            "nuevos": nuevos,
            "en_siigo": en_siigo,
            "fecha_generacion": (root.attrib.get("fecha_generacion") or "")[:19],
        }
    except Exception:
        return None


def sincronizar_historial_desde_importaciones() -> int:
    """
    Rellena el historial con XMLs ya generados en importaciones_productos/.
    Cubre facturas procesadas antes de existir el registro en JSON.
    """
    if not os.path.isdir(CARPETA_IMPORTACIONES):
        return 0

    state = _cargar_historial()
    historial = list(state.get("historial") or [])
    vistos = {
        str(r.get("numero_factura") or "").strip().upper()
        for r in historial
        if r.get("numero_factura")
    }
    agregados = 0

    for fname in sorted(os.listdir(CARPETA_IMPORTACIONES)):
        if not fname.endswith(" codigos siigo.xml"):
            continue
        base = fname[: -len(" codigos siigo.xml")].strip()
        ruta_xml = os.path.join(CARPETA_IMPORTACIONES, fname)
        parsed = _parsear_xml_historial_importacion(ruta_xml)
        if not parsed:
            continue
        numero = parsed["numero_factura"]
        clave = numero.strip().upper()
        if clave in vistos:
            continue

        ruta_excel = os.path.join(CARPETA_IMPORTACIONES, f"{base} registro productos.xlsx")
        sufijo = _sufijo_factura(numero)
        ts = parsed.get("fecha_generacion") or parsed.get("fecha_factura") or datetime.now().isoformat()
        slug_ts = re.sub(r"\W", "", str(ts))[:14] or str(agregados)
        registro = {
            "id": f"imp-{clave}-{slug_ts}",
            "sufijo": sufijo,
            "numero_factura": numero,
            "proveedor": parsed.get("proveedor") or "",
            "nit": parsed.get("nit") or "",
            "total": parsed.get("total") or 0,
            "fecha_factura": parsed.get("fecha_factura") or "",
            "items_count": parsed.get("items_count") or 0,
            "accion": "inventario",
            "estado": "ok",
            "origen": "reconstruido",
            "nuevos": parsed.get("nuevos") or 0,
            "en_siigo": parsed.get("en_siigo") or 0,
            "items_resumen": parsed.get("items_resumen") or [],
            "ruta_excel": os.path.basename(ruta_excel) if os.path.isfile(ruta_excel) else None,
            "ruta_xml": fname,
            "siigo_id": None,
            "mensaje": "Reconstruido desde importaciones_productos/",
            "timestamp": ts if "T" in str(ts) else f"{ts}T12:00:00",
        }
        historial.append(registro)
        vistos.add(clave)
        agregados += 1

    if agregados:
        historial.sort(key=lambda x: x.get("timestamp") or "", reverse=True)
        state["historial"] = historial[:500]
        _guardar_historial(state)
    return agregados


def _reparar_items_resumen_desde_xml(registro: dict) -> list:
    """Completa precios/cantidades desde el XML guardado si faltan en el historial."""
    items = list(registro.get('items_resumen') or [])
    if items and all(_precio_referencia_item(it) for it in items if it.get('codigo')):
        return items
    ruta_xml = registro.get('ruta_xml')
    if not ruta_xml:
        return items
    path = ruta_xml if os.path.isabs(ruta_xml) else os.path.join(CARPETA_IMPORTACIONES, ruta_xml)
    if not os.path.isfile(path):
        return items
    parsed = _parsear_xml_historial_importacion(path)
    if not parsed:
        return items
    return parsed.get('items_resumen') or items


def _calcular_tendencias_precio_historial(historial: list) -> dict:
    """
    Recorre el historial cronológico y calcula tendencia de precio por (registro_id, codigo).
    Compara precio neto unitario contra la compra anterior del mismo código SIIGO.
    """
    ordenados = sorted(historial, key=lambda r: r.get('timestamp') or '')
    ultimo_precio: dict[str, float] = {}
    tendencias: dict[tuple, dict] = {}

    for reg in ordenados:
        if str(reg.get('accion', '')).lower() != 'inventario':
            continue
        rid = reg.get('id')
        items = _reparar_items_resumen_desde_xml(reg)
        for item in items:
            codigo = str(item.get('codigo') or '').strip().upper()
            if not codigo:
                continue
            precio = _precio_referencia_item(item)
            if precio is None:
                continue
            clave = (rid, codigo)
            if codigo in ultimo_precio:
                anterior = ultimo_precio[codigo]
                tend = _tendencia_precio(precio, anterior)
                variacion = round((precio - anterior) / anterior * 100, 2) if anterior else None
                tendencias[clave] = {
                    'precio_anterior': anterior,
                    'tendencia_precio': tend,
                    'variacion_pct': variacion if tend != 'igual' else 0.0,
                }
            else:
                tendencias[clave] = {
                    'precio_anterior': None,
                    'tendencia_precio': 'nuevo',
                    'variacion_pct': None,
                }
            ultimo_precio[codigo] = precio
    return tendencias


def _enriquecer_registro_historial(registro: dict, tendencias: dict) -> dict:
    """Copia el registro con ítems completos y tendencias de precio."""
    reg = dict(registro)
    rid = reg.get('id')
    items = []
    for item in _reparar_items_resumen_desde_xml(reg):
        it = dict(item)
        codigo = str(it.get('codigo') or '').strip().upper()
        extra = tendencias.get((rid, codigo), {})
        it.update(extra)
        items.append(it)
    reg['items_resumen'] = items
    return reg


def _anio_de_registro_historial(row: dict) -> int | None:
    """Año de la factura: fecha_factura → timestamp → id YYYY…"""
    for key in ("fecha_factura", "timestamp", "fecha"):
        raw = str(row.get(key) or "").strip()
        if len(raw) >= 4 and raw[:4].isdigit():
            try:
                return int(raw[:4])
            except ValueError:
                pass
    rid = str(row.get("id") or "")
    if len(rid) >= 4 and rid[:4].isdigit():
        try:
            return int(rid[:4])
        except ValueError:
            pass
    return None


def listar_historial_facturas(
    limit: int = 100,
    accion: str | None = None,
    q: str | None = None,
    anio: int | None = None,
) -> dict:
    """Lista facturas ya procesadas (más recientes primero).

    Por defecto solo el año en curso (fecha de factura; si falta, timestamp).
    """
    sincronizar_historial_desde_importaciones()
    historial = _cargar_historial().get('historial', [])
    tendencias = _calcular_tendencias_precio_historial(historial)
    accion_f = (accion or '').strip().lower()
    q_norm = _normalizar(q or '').strip()
    anio_filtro = int(anio) if anio is not None else int(datetime.now().year)
    filtradas = []
    for row in historial:
        if accion_f and str(row.get('accion', '')).lower() != accion_f:
            continue
        anio_row = _anio_de_registro_historial(row)
        if anio_row is not None and anio_row != anio_filtro:
            continue
        if anio_row is None:
            continue
        enriquecido = _enriquecer_registro_historial(row, tendencias)
        if q_norm:
            partes = [
                enriquecido.get('numero_factura', ''),
                enriquecido.get('proveedor', ''),
                enriquecido.get('nit', ''),
            ]
            for it in enriquecido.get('items_resumen') or []:
                if isinstance(it, dict):
                    partes.append(str(it.get('nombre', '')))
                    partes.append(str(it.get('codigo', '')))
            blob = _normalizar(' '.join(partes))
            if q_norm not in blob:
                continue
        filtradas.append(enriquecido)
    total = len(filtradas)
    limit = max(1, min(int(limit or 100), 500))
    return {
        'historial': filtradas[:limit],
        'total': total,
        'mostrando': min(total, limit),
        'anio': anio_filtro,
    }


def _coincidencias_producto_items(items: list, q_norm: str) -> list:
    """Filtra ítems cuyo nombre o código contienen q_norm (ya normalizado)."""
    hits = []
    for it in items or []:
        if not isinstance(it, dict):
            continue
        nombre = str(it.get('nombre') or it.get('description') or '')
        codigo = str(it.get('codigo') or '')
        blob = _normalizar(f'{nombre} {codigo}')
        if q_norm and q_norm in blob:
            hits.append({
                'nombre': nombre[:160],
                'codigo': codigo,
                'cantidad': it.get('cantidad_min', it.get('quantity')),
                'unidad': it.get('unidad_min') or it.get('unidad') or '',
                'precio_neto': it.get('precio_neto'),
                'precio_unitario': it.get('precio_unitario') or it.get('price'),
                'subtotal': it.get('subtotal'),
            })
    return hits


# Archivo Gmail/local consultable en el panel (Consultar factura).
ANIO_CONSULTA_ARCHIVO_MIN = 2022


def anios_consulta_archivo(hasta: int | None = None) -> list[int]:
    """Años con índice de archivo (2022 … año actual inclusive)."""
    fin = int(hasta if hasta is not None else datetime.now().year)
    if fin < ANIO_CONSULTA_ARCHIVO_MIN:
        return []
    return list(range(ANIO_CONSULTA_ARCHIVO_MIN, fin + 1))


def _ruta_indice_consulta_anio(anio: int) -> str:
    return os.path.join(
        os.path.dirname(__file__), "..", "data", f"facturas_consulta_{int(anio)}.json"
    )


def _cargar_indice_consulta_anio(anio: int) -> dict:
    path = _ruta_indice_consulta_anio(anio)
    try:
        if os.path.isfile(path):
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                data.setdefault("facturas", [])
                return data
    except Exception:
        pass
    return {"anio": int(anio), "facturas": [], "actualizado": None, "fuente": None}


def _guardar_indice_consulta_anio(anio: int, data: dict) -> None:
    path = _ruta_indice_consulta_anio(anio)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _xml_a_registro_consulta(datos: dict, *, origen: str, fuente: str = "") -> dict | None:
    if not datos:
        return None
    numero = f"{datos.get('prefix', '')}{datos.get('number', '')}".strip()
    if not numero:
        return None
    items_raw = datos.get("items") or []
    items_resumen = []
    for it in items_raw:
        if not isinstance(it, dict):
            continue
        items_resumen.append({
            "nombre": it.get("description") or it.get("nombre") or "",
            "codigo": it.get("codigo") or "",
            "cantidad": it.get("quantity"),
            "unidad": it.get("unidad") or "",
            "precio_unitario": it.get("price"),
            "precio_neto": it.get("precio_neto"),
            "subtotal": it.get("subtotal"),
        })
    total = datos.get("total_neto")
    if total is None:
        total = round(sum(float(it.get("subtotal") or 0) for it in items_raw), 2)
    fecha = (datos.get("fecha") or "")[:10]
    return {
        "origen": origen,
        "id": f"{origen}-{numero}",
        "sufijo": _sufijo_factura(numero),
        "numero_factura": numero,
        "proveedor": datos.get("proveedor") or "",
        "nit": datos.get("nit_proveedor") or datos.get("nit") or "",
        "fecha": fecha,
        "total": total or 0,
        "accion": None,
        "estado": "archivo",
        "timestamp": f"{fecha}T12:00:00" if fecha else "",
        "items_resumen": items_resumen,
        "items_count": len(items_resumen),
        "fuente": fuente,
    }


def construir_indices_consulta_rango(
    desde: int = ANIO_CONSULTA_ARCHIVO_MIN,
    hasta: int | None = None,
    forzar: bool = False,
) -> dict:
    """Construye índices de archivo año por año (p. ej. 2022–actual)."""
    anios = [
        a
        for a in anios_consulta_archivo(hasta)
        if a >= int(desde)
    ]
    detalle: list[dict] = []
    total = 0
    for a in anios:
        r = construir_indice_consulta_anio(a, forzar=forzar)
        detalle.append(r)
        total += int(r.get("facturas") or 0)
    return {
        "ok": True,
        "desde": int(desde),
        "hasta": anios[-1] if anios else None,
        "anios": anios,
        "facturas": total,
        "detalle": detalle,
        "mensaje": (
            f"Índices {anios[0]}–{anios[-1]}: {total} factura(s)."
            if anios
            else "Sin años en el rango solicitado."
        ),
    }


def construir_indice_consulta_anio(anio: int = 2025, forzar: bool = False) -> dict:
    """
    Indexa facturas de un año (desde 2022) desde Gmail + ZIPs locales.
    Guarda caché en app/data/facturas_consulta_{anio}.json para consultas rápidas.
    """
    anio = int(anio)
    cache = _cargar_indice_consulta_anio(anio)
    if (
        not forzar
        and cache.get("facturas")
        and cache.get("actualizado")
    ):
        return {
            "ok": True,
            "anio": anio,
            "facturas": len(cache["facturas"]),
            "actualizado": cache.get("actualizado"),
            "cache_hit": True,
            "mensaje": f"Índice {anio} ya cargado ({len(cache['facturas'])} facturas).",
        }

    from app.tools.sincronizar_facturas_de_compra_siigo import (
        GmailAuthError,
        get_gmail_service,
        leer_correos_facturas_periodo,
        descargar_y_extraer_zip,
        extraer_datos_xml_dian,
        CARPETA_FACTURAS_LOCAL,
    )

    fecha_desde = f"{anio}/01/01"
    fecha_hasta = f"{anio + 1}/01/01"
    facturas: list[dict] = []
    vistos: set[str] = set()
    errores: list[str] = []

    # 1) ZIPs / XML locales cuyo IssueDate cae en el año
    try:
        for fname in os.listdir(CARPETA_FACTURAS_LOCAL):
            lower = fname.lower()
            path = os.path.join(CARPETA_FACTURAS_LOCAL, fname)
            xml_content = None
            if lower.endswith(".xml"):
                try:
                    with open(path, encoding="utf-8", errors="ignore") as f:
                        xml_content = f.read()
                except Exception:
                    continue
            elif lower.endswith(".zip"):
                try:
                    import zipfile
                    with zipfile.ZipFile(path, "r") as zf:
                        for name in zf.namelist():
                            if name.lower().endswith(".xml"):
                                xml_content = zf.read(name).decode("utf-8", errors="ignore")
                                break
                except Exception:
                    continue
            if not xml_content:
                continue
            datos = extraer_datos_xml_dian(xml_content)
            if not datos:
                continue
            fecha = (datos.get("fecha") or "")[:10]
            if not fecha.startswith(str(anio)):
                continue
            reg = _xml_a_registro_consulta(datos, origen=f"archivo-{anio}", fuente=fname)
            if not reg:
                continue
            clave = reg["numero_factura"].upper()
            if clave in vistos:
                continue
            vistos.add(clave)
            facturas.append(reg)
    except Exception as e:
        errores.append(f"local: {e}")

    # 2) Gmail del año (incluye ya descargados)
    gmail_ok = False
    try:
        correos = leer_correos_facturas_periodo(
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            solo_no_descargados=False,
        )
        service = get_gmail_service()
        gmail_ok = True
        for correo in correos:
            for adj in correo.get("adjuntos_zip") or []:
                try:
                    xml_content = None
                    zip_path = adj.get("zip_path")
                    if zip_path and os.path.isfile(zip_path):
                        import zipfile
                        with zipfile.ZipFile(zip_path, "r") as zf:
                            for name in zf.namelist():
                                if name.lower().endswith(".xml"):
                                    xml_content = zf.read(name).decode("utf-8", errors="ignore")
                                    break
                    if not xml_content:
                        xml_content, _pdf, _name = descargar_y_extraer_zip(
                            service, correo["id"], adj["id"], adj["filename"]
                        )
                    if not xml_content:
                        continue
                    datos = extraer_datos_xml_dian(xml_content)
                    if not datos:
                        continue
                    fecha = (datos.get("fecha") or "")[:10]
                    if not fecha.startswith(str(anio)):
                        continue
                    reg = _xml_a_registro_consulta(
                        datos,
                        origen=f"gmail-{anio}",
                        fuente=adj.get("filename") or "",
                    )
                    if not reg:
                        continue
                    clave = reg["numero_factura"].upper()
                    if clave in vistos:
                        continue
                    vistos.add(clave)
                    facturas.append(reg)
                except Exception as e:
                    errores.append(f"{adj.get('filename')}: {e}")
    except GmailAuthError as e:
        errores.append(str(e))
    except Exception as e:
        errores.append(f"gmail: {e}")

    facturas.sort(key=lambda r: r.get("fecha") or "", reverse=True)
    payload = {
        "anio": anio,
        "facturas": facturas,
        "actualizado": datetime.now().isoformat(),
        "fuente": "gmail+local" if gmail_ok else "local",
        "errores": errores[:20],
    }
    _guardar_indice_consulta_anio(anio, payload)

    mensaje = f"Índice {anio}: {len(facturas)} factura(s)."
    if not facturas and errores:
        mensaje += " " + (errores[0][:180] if errores else "")
    elif not facturas:
        mensaje += " No se encontraron facturas de ese año en Gmail/local."

    return {
        "ok": True,
        "anio": anio,
        "facturas": len(facturas),
        "actualizado": payload["actualizado"],
        "cache_hit": False,
        "gmail_ok": gmail_ok,
        "errores": errores[:10],
        "mensaje": mensaje,
    }


def estado_indice_consulta_anio(anio: int = 2025) -> dict:
    cache = _cargar_indice_consulta_anio(anio)
    n = len(cache.get("facturas") or [])
    # "listo" = ya se intentó indexar (aunque el año no tenga facturas).
    listo = bool(cache.get("actualizado")) or n > 0
    return {
        "anio": int(anio),
        "facturas": n,
        "actualizado": cache.get("actualizado"),
        "fuente": cache.get("fuente"),
        "listo": listo,
    }


def _filtro_anio_fecha(fecha: str, anio: int | None) -> bool:
    if anio is None:
        return True
    return (fecha or "").startswith(str(anio))


def consultar_facturas_por_producto(
    q: str,
    limit: int = 50,
    anio: int | None = None,
    *,
    asegurar_indice_anio: bool = True,
) -> dict:
    """
    Busca facturas de proveedores (pendientes + historial + archivo por año)
    que contengan un producto por nombre o código.

    anio=2022..actual incluye el índice Gmail/local de ese año (se construye si falta).
    anio=None busca en todas las fuentes disponibles (archivo desde 2022).
    """
    q_limpia = (q or '').strip()
    q_norm = _normalizar(q_limpia)
    if len(q_norm) < 2:
        return {
            'ok': True,
            'q': q_limpia,
            'anio': anio,
            'resultados': [],
            'total': 0,
            'mensaje': 'Escribe al menos 2 caracteres para buscar por producto.',
        }

    sincronizar_historial_desde_importaciones()
    resultados: list[dict] = []
    avisos: list[str] = []

    # ── Pendientes (correo en cola) ──────────────────────────────────────────
    try:
        pendientes = _cargar_pendientes().get('pendientes', {}) or {}
    except Exception:
        pendientes = {}
    for sufijo, entrada in pendientes.items():
        if not isinstance(entrada, dict):
            continue
        try:
            datos = json.loads(entrada.get('datos_json') or '{}')
        except Exception:
            datos = {}
        fecha = (datos.get('fecha') or '')[:10]
        if not _filtro_anio_fecha(fecha, anio):
            continue
        items_raw = datos.get('items') or []
        hits = _coincidencias_producto_items(items_raw, q_norm)
        if not hits:
            continue
        resultados.append({
            'origen': 'pendiente',
            'id': f'pendiente-{sufijo}',
            'sufijo': str(sufijo).upper(),
            'numero_factura': entrada.get('numero_factura') or datos.get('number') or '',
            'proveedor': entrada.get('proveedor') or datos.get('proveedor') or '',
            'nit': entrada.get('nit') or datos.get('nit') or '',
            'fecha': fecha,
            'total': entrada.get('total') or datos.get('total_neto') or 0,
            'accion': None,
            'estado': entrada.get('estado') or 'pendiente',
            'timestamp': entrada.get('timestamp') or '',
            'coincidencias': hits,
            'items_count': entrada.get('items_count') or len(items_raw),
        })

    # ── Historial (ya procesadas desde el correo) ────────────────────────────
    historial = _cargar_historial().get('historial', [])
    tendencias = _calcular_tendencias_precio_historial(historial)
    for row in historial:
        enriquecido = _enriquecer_registro_historial(row, tendencias)
        fecha = (enriquecido.get('fecha_factura') or '')[:10]
        if not _filtro_anio_fecha(fecha, anio):
            continue
        hits = _coincidencias_producto_items(enriquecido.get('items_resumen') or [], q_norm)
        if not hits:
            continue
        resultados.append({
            'origen': 'historial',
            'id': enriquecido.get('id') or '',
            'sufijo': enriquecido.get('sufijo') or '',
            'numero_factura': enriquecido.get('numero_factura') or '',
            'proveedor': enriquecido.get('proveedor') or '',
            'nit': enriquecido.get('nit') or '',
            'fecha': fecha,
            'total': enriquecido.get('total') or 0,
            'accion': enriquecido.get('accion'),
            'estado': enriquecido.get('estado') or '',
            'timestamp': enriquecido.get('timestamp') or '',
            'coincidencias': hits,
            'items_count': enriquecido.get('items_count') or len(enriquecido.get('items_resumen') or []),
        })

    # ── Archivo por año (Gmail/local, desde 2022) ────────────────────────────
    anio_hoy = datetime.now().year
    anios_archivo: list[int] = []
    if anio is None:
        anios_archivo = anios_consulta_archivo(anio_hoy)
    elif ANIO_CONSULTA_ARCHIVO_MIN <= int(anio) <= anio_hoy:
        anios_archivo = [int(anio)]

    anios_sin_indice: list[int] = []
    for anio_arch in anios_archivo:
        cache = _cargar_indice_consulta_anio(anio_arch)
        indexado = bool(cache.get("actualizado")) or bool(cache.get("facturas"))
        if asegurar_indice_anio and anio == anio_arch and not indexado:
            try:
                build = construir_indice_consulta_anio(anio_arch, forzar=False)
                if build.get("mensaje"):
                    avisos.append(str(build["mensaje"]))
                cache = _cargar_indice_consulta_anio(anio_arch)
                indexado = bool(cache.get("actualizado")) or bool(cache.get("facturas"))
            except Exception as e:
                avisos.append(f"No se pudo cargar archivo {anio_arch}: {e}")
        elif not indexado:
            anios_sin_indice.append(anio_arch)
        for row in cache.get("facturas") or []:
            fecha = (row.get("fecha") or "")[:10]
            if not _filtro_anio_fecha(fecha, anio):
                continue
            # Evitar duplicar si ya está en historial/pendientes
            num = (row.get("numero_factura") or "").upper()
            if any((r.get("numero_factura") or "").upper() == num for r in resultados):
                continue
            hits = _coincidencias_producto_items(row.get("items_resumen") or [], q_norm)
            if not hits:
                continue
            resultados.append({
                **row,
                "coincidencias": hits,
            })

    if anios_sin_indice:
        if len(anios_sin_indice) == 1:
            avisos.append(
                f"Aún no hay índice de archivo {anios_sin_indice[0]}. "
                "Se está indexando o pulsa «Cargar archivo»."
            )
        else:
            rango = f"{anios_sin_indice[0]}–{anios_sin_indice[-1]}"
            avisos.append(
                f"Falta indexar archivo {rango}. "
                f"Pulsa «Cargar archivo» una vez (desde {ANIO_CONSULTA_ARCHIVO_MIN})."
            )

    # Pendientes primero; resto por fecha/timestamp descendente
    pendientes_r = [r for r in resultados if r.get('origen') == 'pendiente']
    otros_r = sorted(
        [r for r in resultados if r.get('origen') != 'pendiente'],
        key=lambda r: r.get('timestamp') or r.get('fecha') or '',
        reverse=True,
    )
    resultados = pendientes_r + otros_r

    limit = max(1, min(int(limit or 50), 200))
    total = len(resultados)
    indices_keys = anios_archivo or anios_consulta_archivo(anio_hoy)
    return {
        'ok': True,
        'q': q_limpia,
        'anio': anio,
        'resultados': resultados[:limit],
        'total': total,
        'mostrando': min(total, limit),
        'avisos': avisos,
        'anio_min': ANIO_CONSULTA_ARCHIVO_MIN,
        'indices': {
            str(a): estado_indice_consulta_anio(a) for a in indices_keys
        },
    }


def obtener_registro_historial(registro_id: str) -> dict | None:
    rid = (registro_id or '').strip()
    if not rid:
        return None
    historial = _cargar_historial().get('historial', [])
    tendencias = _calcular_tendencias_precio_historial(historial)
    for row in historial:
        if row.get('id') == rid:
            return _enriquecer_registro_historial(row, tendencias)
    return None


# ─────────────────────────────────────────────
#  Motor de procesamiento (interno)
# ─────────────────────────────────────────────

def _ejecutar_procesamiento(numero_factura: str, datos: dict, xml_content: str, silent: bool = False) -> dict:
    """
    Extrae productos, genera Excel + XML.
    silent=True: omite envíos por WhatsApp (modo terminal).
    Retorna un dict con los archivos generados y el resumen.
    """
    proveedor = datos.get('proveedor', '')
    productos_nuevos = []
    productos_duplicados = []
    codigos_en_factura = set()   # evita colisiones dentro de la misma factura

    for item in datos.get('items', []):
        nombre = item.get('description', '').strip()
        subtotal = item.get('subtotal', 0)
        cantidad_original = item.get('quantity', 1)

        iva_linea = sum(
            imp['valor'] for imp in item.get('impuestos', [])
            if imp.get('id_dian') == '01'
        )
        unit_code = _extraer_unit_code_de_xml(xml_content, nombre)
        cantidad_min, unidad_min, codigo_dian_min = convertir_a_unidad_minima(
            cantidad_original, unit_code
        )

        # Multiplicador de contenido: "caja x 100 unidades" → factor 100
        # Solo se aplica si no hay volumen líquido en la descripción
        multiplicador = _extraer_multiplicador_descripcion(nombre)
        if multiplicador > 1:
            cantidad_min = round(cantidad_min * multiplicador, 6)
            # La unidad mínima siempre pasa a ser 'Un' cuando hay multiplicador en descripción
            unidad_min      = 'Un'
            codigo_dian_min = 'NAR'
            print(f"  📦 Contenido detectado: {int(cantidad_original)} × {multiplicador} = {int(cantidad_min)} {unidad_min}")
        elif unidad_min == 'Un':
            # Si la unidad del proveedor es UN pero el producto tiene volumen o masa
            # indicados en la descripción, convertir a la unidad mínima real
            vol_ml = _extraer_volumen_ml_descripcion(nombre)
            if vol_ml > 0 and not _descripcion_es_envase_unitario(nombre):
                cantidad_min    = round(cantidad_min * vol_ml, 6)
                unidad_min      = 'mL'
                codigo_dian_min = 'MLT'
                print(f"  💧 Líquido detectado: {int(cantidad_original)} × {vol_ml} mL = {cantidad_min:.0f} mL")
            elif vol_ml > 0:
                print(f"  🧴 Envase detectado: {nombre[:40]} → se conserva como unidad")
            else:
                masa_g = _extraer_masa_g_descripcion(nombre)
                if masa_g > 0:
                    cantidad_min    = round(cantidad_min * masa_g, 6)
                    unidad_min      = 'g'
                    codigo_dian_min = 'GRM'
                    print(f"  ⚖️  Sólido detectado: {int(cantidad_original)} × {masa_g} g = {cantidad_min:.0f} g")

        precio_unitario = calcular_precio_unitario_min(subtotal, iva_linea, cantidad_min)
        # Precio neto (sin IVA) — usado en ítems de compra SIIGO para que el IVA se aplique correctamente
        # Usamos 6 decimales para evitar discrepancias de redondeo al multiplicar por la cantidad
        precio_neto = round(subtotal / cantidad_min, 6) if cantidad_min > 0 else 0.0

        # Genera código único dentro de esta factura (evita GOTPIP77MUn × 2 en misma factura)
        referencia_proveedor = _extraer_referencia_proveedor_de_xml(xml_content, nombre)
        codigo_ref = _codigo_siigo_desde_referencia(referencia_proveedor)
        if codigo_ref:
            codigo = codigo_ref
            print(f"  🔗 Referencia SIIGO: {codigo} — {nombre[:40]}")
        else:
            codigo = generar_codigo_producto(nombre, unidad_min, codigos_en_factura)
        codigos_en_factura.add(codigo)
        existe_en_siigo = verificar_producto_en_siigo(codigo)

        producto = {
            'nombre':            nombre,
            'codigo':            codigo,
            'cantidad_original': cantidad_original,
            'unidad_original':   unit_code,
            'multiplicador':     multiplicador,
            'cantidad_min':      cantidad_min,
            'unidad_min':        unidad_min,
            'codigo_dian_min':   codigo_dian_min,
            'subtotal':          subtotal,
            'iva':               iva_linea,
            'precio_unitario':   precio_unitario,  # con IVA — para Excel/precio venta
            'precio_neto':       precio_neto,       # sin IVA — para ítems de compra SIIGO
            'existe_en_siigo':   existe_en_siigo,
            'duplicado':         existe_en_siigo,
        }
        if existe_en_siigo:
            productos_duplicados.append(producto)
            print(f"  📦 En SIIGO: {codigo} — {nombre[:40]} (suma inventario)")
        else:
            productos_nuevos.append(producto)
            print(f"  ✅ Nuevo: {codigo} — {nombre[:40]} → ${precio_unitario:.2f}/{unidad_min}")

    todos = productos_nuevos + productos_duplicados
    if not todos:
        print(f"  ℹ️ Sin ítems procesables en {numero_factura}.")
        return {}

    ruta_excel = generar_excel_importacion(todos, numero_factura)
    ruta_xml   = generar_xml_compra_siigo(datos, todos, numero_factura)

    arch = {
        'ruta':           ruta_excel,
        'ruta_xml':       ruta_xml,
        'numero_factura': numero_factura,
        'proveedor':      proveedor,
        'nuevos':         len(productos_nuevos),
        'duplicados':     len(productos_duplicados),
        'productos':      todos,   # lista completa para flujo API
    }

    if not silent:
        enviar_whatsapp_reporte(_construir_resumen_whatsapp(arch), numero_destino=GRUPO_COMPRAS)
        if arch['ruta']:
            enviar_whatsapp_archivo(
                arch['ruta'],
                f"📊 *Excel productos SIIGO* — Factura {numero_factura}",
                numero_destino=GRUPO_COMPRAS,
            )
        if ruta_xml and os.path.exists(ruta_xml):
            enviar_whatsapp_archivo(
                ruta_xml,
                (
                    f"📄 *XML de compra SIIGO* — Factura {numero_factura}\n"
                    f"Usa: Compras → *Crear compra o gasto desde un XML o ZIP*"
                ),
                numero_destino=GRUPO_COMPRAS,
            )
    return arch


# ─────────────────────────────────────────────
#  Notificación de la siguiente factura en cola
# ─────────────────────────────────────────────

def _notificar_siguiente_factura_pendiente():
    """
    Envía al grupo la notificación de la primera factura pendiente en la cola.
    Se llama después de encolar nuevas facturas o tras procesar una existente,
    para mantener el flujo de una factura a la vez.
    """
    state = _cargar_pendientes()
    pendientes = state.get('pendientes', {})
    if not pendientes:
        return

    # Tomar la primera entrada (orden de inserción, Python 3.7+)
    sufijo, entrada = next(iter(pendientes.items()))
    numero_factura = entrada['numero_factura']
    proveedor      = entrada['proveedor']
    nit            = entrada.get('nit', '')
    n_items        = entrada.get('items_count', 0)
    total          = entrada.get('total', 0)
    es_nuevo       = entrada.get('es_nuevo_proveedor', False)

    estado_prov = "⚠️ Proveedor nuevo" if es_nuevo else "✅ Proveedor conocido"
    pendientes_restantes = len(pendientes)
    cabecera = (
        f"─────────────────────────\n"
        f"⏭️ *Siguiente en cola ({pendientes_restantes - 1} más después):*\n\n"
        if pendientes_restantes > 1 else ""
    )
    msg = (
        f"{cabecera}"
        f"📦 *Nueva factura de compra*  —  código: `{sufijo}`\n\n"
        f"🔢 {numero_factura}\n"
        f"🏢 {proveedor}  |  NIT: {nit or '—'}\n"
        f"📦 {n_items} ítem(s)  |  💰 ${total:,.0f} COP\n"
        f"{estado_prov}\n\n"
        f"👉 Clasifícala en el *Panel de Operaciones*\n"
        f"   _(o usa *inv gasto {sufijo}* / *inv skip {sufijo}* desde aquí)_"
    )

    enviar_whatsapp_reporte(msg, numero_destino=GRUPO_COMPRAS)
    print(f"  ✉️  Notificación enviada al grupo — código: {sufijo}")


# ─────────────────────────────────────────────
#  Panel de operaciones — inspección y proceso manual
# ─────────────────────────────────────────────

def _codigo_manual_valido(codigo: str) -> str:
    codigo = _sanitizar_codigo_siigo(codigo)
    if not codigo:
        return ''
    if not re.match(r'^[A-Za-z0-9._-]{2,40}$', codigo):
        raise ValueError(f"Código SIIGO inválido: {codigo}")
    return codigo


def _computar_items_factura(datos: dict, xml_content: str, codigos_manual: dict | None = None) -> list:
    """
    Computa ítems con códigos McKenna, unidades y precios SIN generar archivos.
    Para mostrar en el panel antes de que el usuario decida qué incluir.
    """
    items_out = []
    codigos_en_factura = set()
    codigos_manual = codigos_manual or {}
    for idx, item in enumerate(datos.get('items', [])):
        nombre = item.get('description', '').strip()
        subtotal = item.get('subtotal', 0)
        cantidad_original = item.get('quantity', 1)
        iva_linea = sum(
            imp['valor'] for imp in item.get('impuestos', [])
            if imp.get('id_dian') == '01'
        )
        unit_code = _extraer_unit_code_de_xml(xml_content, nombre)
        cantidad_min, unidad_min, codigo_dian_min = convertir_a_unidad_minima(
            cantidad_original, unit_code
        )
        multiplicador = _extraer_multiplicador_descripcion(nombre)
        if multiplicador > 1:
            cantidad_min    = round(cantidad_min * multiplicador, 6)
            unidad_min      = 'Un'
            codigo_dian_min = 'NAR'
        elif unidad_min == 'Un':
            vol_ml = _extraer_volumen_ml_descripcion(nombre)
            if vol_ml > 0 and not _descripcion_es_envase_unitario(nombre):
                cantidad_min    = round(cantidad_min * vol_ml, 6)
                unidad_min      = 'mL'
                codigo_dian_min = 'MLT'
            elif vol_ml > 0:
                pass
            else:
                masa_g = _extraer_masa_g_descripcion(nombre)
                if masa_g > 0:
                    cantidad_min    = round(cantidad_min * masa_g, 6)
                    unidad_min      = 'g'
                    codigo_dian_min = 'GRM'
        precio_unitario = calcular_precio_unitario_min(subtotal, iva_linea, cantidad_min)
        precio_neto     = round(subtotal / cantidad_min, 6) if cantidad_min > 0 else 0.0
        referencia_proveedor = _extraer_referencia_proveedor_de_xml(xml_content, nombre)
        codigo_manual = _codigo_manual_valido(
            str(codigos_manual.get(str(idx)) or codigos_manual.get(idx) or '')
        )
        codigo_por_referencia = False
        if codigo_manual:
            codigo = codigo_manual
        else:
            codigo_ref = _codigo_siigo_desde_referencia(referencia_proveedor)
            if codigo_ref:
                codigo = codigo_ref
                codigo_por_referencia = True
            else:
                codigo = generar_codigo_producto(nombre, unidad_min, codigos_en_factura)
        codigos_en_factura.add(codigo)
        producto_siigo  = buscar_producto_en_siigo_por_codigo(codigo)
        existe_en_siigo = producto_siigo is not None
        items_out.append({
            'indice':            idx,
            'nombre':            nombre,
            'codigo':            codigo,
            'codigo_sugerido':   generar_codigo_producto(nombre, unidad_min, set()),
            'codigo_manual':     bool(codigo_manual),
            'codigo_por_referencia': codigo_por_referencia,
            'referencia_proveedor': referencia_proveedor,
            'cantidad_original': cantidad_original,
            'unidad_original':   unit_code,
            'multiplicador':     multiplicador,
            'cantidad_min':      round(cantidad_min, 4),
            'unidad_min':        unidad_min,
            'codigo_dian_min':   codigo_dian_min,
            'subtotal':          subtotal,
            'iva':               iva_linea,
            'precio_unitario':   precio_unitario,
            'precio_neto':       precio_neto,
            'existe_en_siigo':   existe_en_siigo,
            'duplicado':         existe_en_siigo,
            'siigo_producto':    _resumen_producto_siigo(producto_siigo),
            'impuestos':         item.get('impuestos', []),
            'precio_proveedor':  item.get('price', 0),
        })
    return items_out


def obtener_detalle_factura(sufijo: str) -> dict | None:
    """Retorna factura completa con ítems computados para el panel."""
    key, entrada = _buscar_pendiente(sufijo)
    if not entrada:
        return None
    datos = json.loads(entrada['datos_json'])
    xml_content = base64.b64decode(entrada['xml_b64']).decode('utf-8')
    items = _computar_items_factura(datos, xml_content)
    compra_registrada = buscar_compra_siigo_registrada(
        datos,
        obtener_compras_siigo_para_dedupe(),
    )
    return {
        'sufijo':             sufijo,
        'numero_factura':     entrada['numero_factura'],
        'proveedor':          entrada['proveedor'],
        'nit':                entrada.get('nit', ''),
        'es_nuevo_proveedor': entrada.get('es_nuevo_proveedor', False),
        'total':              entrada.get('total', 0),
        'estado':             entrada.get('estado', ''),
        'fecha':              datos.get('fecha', ''),
        'total_bruto':        datos.get('total_bruto', 0),
        'total_descuentos':   datos.get('total_descuentos', 0),
        'total_neto':         datos.get('total_neto', 0),
        'compra_registrada_siigo': compra_registrada,
        'items':              items,
        'timestamp':          entrada.get('timestamp', ''),
    }


def revisar_codigo_producto_siigo(codigo: str) -> dict:
    """Valida un código manual y consulta si ya existe en SIIGO."""
    codigo_limpio = _codigo_manual_valido(codigo)
    producto = buscar_producto_en_siigo_por_codigo(codigo_limpio)
    existe = producto is not None
    return {
        'codigo': codigo_limpio,
        'existe_en_siigo': existe,
        'duplicado': existe,
        'siigo_producto': _resumen_producto_siigo(producto),
    }


_SIIGO_UNIT_API = {'Un': '94', 'mL': '79', 'g': '62'}


def crear_producto_en_siigo(producto: dict) -> dict:
    """
    Crea un producto inventariable en SIIGO vía POST /v1/products.
    Retorna {ok, mensaje|error, siigo_producto?}.

    Campos útiles en `producto`:
      codigo, nombre, unidad_min, precio_neto (costo),
      precio_unitario (base compra/venta según flujo factura),
      precio_lista (opcional: precio de lista final; si viene, no se aplica ×1.3),
      iva (>0 → impuesto 3118).
    """
    codigo = _codigo_manual_valido(str(producto.get('codigo', '')).strip())
    if not codigo:
        return {'ok': False, 'error': 'Código SIIGO vacío o inválido'}
    existente = buscar_producto_en_siigo_por_codigo(codigo)
    if existente:
        return {
            'ok': False,
            'error': f'El código {codigo} ya existe en SIIGO',
            'siigo_producto': _resumen_producto_siigo(existente),
        }

    token = autenticar_siigo()
    if not token:
        return {'ok': False, 'error': 'No se pudo autenticar con SIIGO'}

    has_iva = float(producto.get('iva') or 0) > 0
    taxes = [{'id': 3118}] if has_iva else []
    try:
        precio_vu = float(producto.get('precio_unitario') or 0)
    except (TypeError, ValueError):
        precio_vu = 0.0
    try:
        precio_neto = float(producto.get('precio_neto') if producto.get('precio_neto') is not None else precio_vu)
    except (TypeError, ValueError):
        precio_neto = precio_vu

    # Precio de lista opcional. Si se envía `prices` con value 0/null, Siigo falla
    # (parameter_required). Sin precio: omitir el bloque completo.
    valor_lista = None
    if 'precio_lista' in producto:
        try:
            pl = producto.get('precio_lista')
            if pl not in (None, '') and float(pl) > 0:
                valor_lista = round(float(pl), 0)
        except (TypeError, ValueError):
            valor_lista = None
    elif precio_vu > 0:
        # Flujo factura (sin precio_lista explícito): deriva lista ≈ ×1.3
        valor_lista = round(precio_vu * 1.3, 0)

    unit_cost = max(0.0, float(precio_neto or 0))
    siigo_unit_code = _SIIGO_UNIT_API.get(producto.get('unidad_min', 'Un'), '94')

    payload = {
        'code': codigo,
        'name': str(producto.get('nombre', ''))[:120],
        'account_group': 297,
        'type': 'Product',
        'stock_control': True,
        'unit': {'code': siigo_unit_code},
        'warehouses': [{'id': 41, 'quantity': 0, 'unit_cost': unit_cost}],
        'taxes': taxes,
    }
    if valor_lista is not None and valor_lista > 0:
        payload['prices'] = [{
            'currency_code': 'COP',
            'price_list': [{'position': 1, 'value': valor_lista}],
        }]
    headers = {
        'Authorization': f'Bearer {token}',
        'Partner-Id': PARTNER_ID,
        'Content-Type': 'application/json',
    }
    try:
        r = requests.post(
            'https://api.siigo.com/v1/products',
            json=payload,
            headers=headers,
            timeout=20,
        )
        if r.status_code in (200, 201):
            data = r.json()
            resumen = _resumen_producto_siigo(data) or {
                'codigo': data.get('code', codigo),
                'nombre': data.get('name', producto.get('nombre', '')),
                'unidad': producto.get('unidad_min', ''),
                'activo': True,
            }
            try:
                from app.services.rentabilidad import registrar_producto_en_cache_costos
                registrar_producto_en_cache_costos(
                    resumen.get('codigo') or codigo,
                    resumen.get('nombre') or str(producto.get('nombre', '')),
                    unit_cost=unit_cost,
                    precio_lista=float(valor_lista or 0),
                )
            except Exception:
                pass
            return {
                'ok': True,
                'mensaje': f"Producto {codigo} creado en SIIGO",
                'siigo_id': data.get('id'),
                'siigo_producto': resumen,
            }
        return {'ok': False, 'error': f'SIIGO HTTP {r.status_code}: {(r.text or "")[:250]}'}
    except Exception as e:
        return {'ok': False, 'error': str(e)}


def crear_productos_factura_en_siigo(
    sufijo: str,
    indices: list,
    codigos_manual: dict | None = None,
) -> dict:
    """Crea en SIIGO los ítems nuevos indicados de una factura pendiente."""
    key, entrada = _buscar_pendiente(sufijo)
    if not entrada:
        return {'ok': False, 'error': f'Factura {sufijo} no encontrada'}

    xml_content = base64.b64decode(entrada['xml_b64']).decode('utf-8')
    datos = json.loads(entrada['datos_json'])
    items = _computar_items_factura(datos, xml_content, codigos_manual)
    indices_norm = sorted({int(i) for i in indices if str(i).isdigit() or isinstance(i, int)})
    if not indices_norm:
        return {'ok': False, 'error': 'Ningún ítem indicado'}

    creados = []
    omitidos = []
    errores = []
    for indice in indices_norm:
        item = next((p for p in items if p.get('indice') == indice), None)
        if not item:
            errores.append({'indice': indice, 'error': 'Ítem no encontrado'})
            continue
        if item.get('existe_en_siigo', item.get('duplicado')):
            omitidos.append({
                'indice': indice,
                'codigo': item.get('codigo'),
                'motivo': 'Ya existe en SIIGO',
                'siigo_producto': item.get('siigo_producto'),
            })
            continue
        resultado = crear_producto_en_siigo(item)
        if resultado.get('ok'):
            creados.append({
                'indice': indice,
                'codigo': item.get('codigo'),
                'nombre': item.get('nombre'),
                'siigo_producto': resultado.get('siigo_producto'),
            })
        else:
            errores.append({
                'indice': indice,
                'codigo': item.get('codigo'),
                'error': resultado.get('error', 'Error desconocido'),
            })

    ok = bool(creados) and not errores
    parcial = bool(creados) and bool(errores)
    return {
        'ok': ok or parcial,
        'parcial': parcial,
        'creados': creados,
        'omitidos': omitidos,
        'errores': errores,
        'mensaje': (
            f"{len(creados)} producto(s) creado(s) en SIIGO"
            + (f", {len(errores)} error(es)" if errores else '')
            + (f", {len(omitidos)} ya existían" if omitidos else '')
        ),
    }


def procesar_items_inventario(sufijo: str, indices: list, codigos_manual: dict | None = None) -> dict:
    """
    Genera Excel + XML solo con los ítems en `indices`.
    Envía reporte de texto al grupo WA (sin adjuntos).
    Quita la factura de la cola.
    """
    key, entrada = _buscar_pendiente(sufijo)
    if not entrada:
        return {'ok': False, 'error': f'Factura {sufijo} no encontrada'}
    xml_content   = base64.b64decode(entrada['xml_b64']).decode('utf-8')
    datos         = json.loads(entrada['datos_json'])

    compra_registrada = buscar_compra_siigo_registrada(datos, obtener_compras_siigo_para_dedupe())
    if compra_registrada:
        doc = compra_registrada.get('name') or compra_registrada.get('id') or 'SIIGO'
        return {
            'ok': False,
            'error': f'Factura ya registrada en SIIGO ({doc}). No se debe inventariar de nuevo.',
            'compra_registrada_siigo': compra_registrada,
        }

    productos_calc = _computar_items_factura(datos, xml_content, codigos_manual)
    indices_norm = sorted({int(i) for i in indices if str(i).isdigit() or isinstance(i, int)})
    items_sel     = [p for p in productos_calc if p.get('indice') in indices_norm]
    if not items_sel:
        return {'ok': False, 'error': 'Ningún ítem seleccionado'}

    ruta_excel = generar_excel_importacion(items_sel, entrada['numero_factura'])
    ruta_xml   = generar_xml_compra_siigo(datos, items_sel, entrada['numero_factura'])
    nuevos = sum(1 for p in items_sel if not p.get('existe_en_siigo', p.get('duplicado')))
    en_siigo = sum(1 for p in items_sel if p.get('existe_en_siigo', p.get('duplicado')))
    excel_linea = (
        f"   • {os.path.basename(ruta_excel)}\n" if ruta_excel else
        "   • _(sin Excel — todos los ítems ya existen en SIIGO)_\n"
    )
    msg = (
        f"✅ *Factura procesada desde el panel*\n\n"
        f"🔢 {entrada['numero_factura']}\n"
        f"🏢 {entrada['proveedor']}\n"
        f"📦 {nuevos} producto(s) nuevo(s)  ·  {en_siigo} en SIIGO (suman inventario)\n"
        f"📎 Archivos generados en el servidor:\n"
        f"{excel_linea}"
        f"   • {os.path.basename(ruta_xml or '—')}"
    )
    enviar_whatsapp_reporte(msg, numero_destino=GRUPO_COMPRAS)
    registrar_factura_historial(
        entrada,
        'inventario',
        sufijo=key,
        origen='panel',
        datos=datos,
        items_resumen=_resumen_items_historial(items_sel),
        nuevos=nuevos,
        en_siigo=en_siigo,
        ruta_excel=ruta_excel,
        ruta_xml=ruta_xml,
        mensaje=msg.replace('*', ''),
    )
    _quitar_pendiente(key)
    return {
        'ok':      True,
        'nuevos':  nuevos,
        'en_siigo': en_siigo,
        'duplicados': en_siigo,
        'ruta_excel': os.path.basename(ruta_excel) if ruta_excel else None,
        'ruta_xml':   os.path.basename(ruta_xml or ''),
        'mensaje': msg,
    }


FECHA_INICIO_COMPRAS_SIIGO = os.getenv("FECHA_INICIO_COMPRAS_SIIGO", "2026-01-01")


def _nit_base(nit: str) -> str:
    """NIT sin dígito de verificación ni separadores."""
    raw = str(nit or '').strip()
    if '-' in raw:
        raw = raw.split('-', 1)[0]
    return re.sub(r'\D', '', raw)


def _factura_id_normalizado(valor: str) -> str:
    return re.sub(r'[^A-Z0-9]', '', (valor or '').upper())


def _variantes_factura(prefix: str, number: str) -> set[str]:
    pref = _factura_id_normalizado(prefix)
    num = _factura_id_normalizado(number)
    variantes = {v for v in (num, f"{pref}{num}") if v}
    if num:
        num_sin_ceros = num.lstrip('0') or '0'
        variantes.add(num_sin_ceros)
        if pref:
            variantes.add(f"{pref}{num_sin_ceros}")
    return variantes


def _variantes_compra_siigo(compra: dict) -> set[str]:
    pi = compra.get('provider_invoice') or {}
    variantes = _variantes_factura(str(pi.get('prefix', '')), str(pi.get('number', '')))
    for campo in ('name', 'number'):
        valor = _factura_id_normalizado(str(compra.get(campo, '')))
        if valor:
            variantes.add(valor)
    return variantes


def _nit_compra_siigo(compra: dict) -> str:
    supplier = compra.get('supplier') or compra.get('provider') or {}
    if isinstance(supplier, dict):
        return _nit_base(str(
            supplier.get('identification')
            or supplier.get('nit')
            or supplier.get('id')
            or ''
        ))
    return ''


def _fecha_compra_siigo(compra: dict) -> str:
    for campo in ('date', 'created', 'created_at', 'updated_at'):
        valor = str(compra.get(campo) or '').strip()
        if valor:
            return valor[:10]
    return ''


def _float_or_none(valor) -> float | None:
    try:
        if valor is None:
            return None
        if isinstance(valor, str):
            valor = valor.replace(',', '').strip()
        return float(valor)
    except (TypeError, ValueError):
        return None


def _valor_compra_siigo(compra: dict) -> float | None:
    for campo in ('total', 'total_value', 'value', 'paid_value'):
        val = _float_or_none(compra.get(campo))
        if val is not None:
            return val
    payments = compra.get('payments')
    if isinstance(payments, list):
        total = 0.0
        found = False
        for p in payments:
            if isinstance(p, dict):
                val = _float_or_none(p.get('value') or p.get('payment_value'))
                if val is not None:
                    total += val
                    found = True
        if found:
            return total
    return None


def _valor_factura_datos(datos: dict) -> float:
    val = _float_or_none(datos.get('total_neto'))
    if val is not None and val > 0:
        return val
    return round(sum(
        item.get('subtotal', 0) + sum(
            imp.get('valor', 0) for imp in item.get('impuestos', [])
        )
        for item in datos.get('items', [])
    ), 2)


def obtener_compras_siigo_para_dedupe(fecha_inicio: str = FECHA_INICIO_COMPRAS_SIIGO) -> list[dict]:
    """Carga compras SIIGO para evitar re-encolar facturas ya registradas."""
    token = autenticar_siigo()
    if not token:
        print("  ⚠️ [SIIGO] Sin token — no se podrán detectar compras ya registradas")
        return []

    headers = {"Authorization": f"Bearer {token}", "Partner-Id": PARTNER_ID}
    compras = []
    pagina = 1
    while True:
        try:
            res = requests.get(
                "https://api.siigo.com/v1/purchases",
                params={"date_start": fecha_inicio, "page": pagina, "page_size": 100},
                headers=headers,
                timeout=20,
            )
            if res.status_code != 200:
                print(f"  ⚠️ [SIIGO] No se pudo consultar compras: HTTP {res.status_code}")
                break
            data = res.json()
            resultados = data.get('results') or []
            compras.extend(resultados)
            pag = data.get('pagination') or {}
            total = int(pag.get('total_results') or len(compras))
            if not resultados or len(compras) >= total:
                break
            pagina += 1
        except Exception as e:
            print(f"  ⚠️ [SIIGO] Error consultando compras registradas: {e}")
            break
    print(f"  🗂️  {len(compras)} compra(s) SIIGO cargadas para detectar duplicados")
    return compras


def buscar_compra_siigo_registrada(datos: dict, compras_siigo: list[dict]) -> dict | None:
    """
    Retorna resumen de compra SIIGO si coincide por número, fecha y valor.
    Si SIIGO no trae alguno de esos datos, usa NIT/proveedor como apoyo.
    Soporta SIIGO guardando prefix/number como ("FE", "32480"), ("", "FE32480")
    o con ceros a la izquierda.
    """
    variantes_xml = _variantes_factura(str(datos.get('prefix', '')), str(datos.get('number', '')))
    nit_xml = _nit_base(str(datos.get('nit_proveedor') or datos.get('nit') or ''))
    fecha_xml = str(datos.get('fecha') or '')[:10]
    valor_xml = _valor_factura_datos(datos)
    for compra in compras_siigo:
        if not (variantes_xml & _variantes_compra_siigo(compra)):
            continue
        nit_siigo = _nit_compra_siigo(compra)
        if nit_xml and nit_siigo and nit_xml != nit_siigo:
            continue

        fecha_siigo = _fecha_compra_siigo(compra)
        if fecha_xml and fecha_siigo and fecha_xml != fecha_siigo:
            continue

        valor_siigo = _valor_compra_siigo(compra)
        if valor_siigo is not None and valor_xml is not None:
            tolerancia = max(2.0, round(abs(valor_xml) * 0.001, 2))
            if abs(valor_siigo - valor_xml) > tolerancia:
                continue

        return {
            'id': compra.get('id', ''),
            'name': compra.get('name', ''),
            'nit': nit_siigo,
            'fecha': fecha_siigo,
            'valor': valor_siigo,
            'match': {
                'numero': True,
                'fecha': bool(fecha_xml and fecha_siigo and fecha_xml == fecha_siigo),
                'valor': valor_siigo is not None and abs(valor_siigo - valor_xml) <= max(2.0, round(abs(valor_xml) * 0.001, 2)),
            },
            'provider_invoice': compra.get('provider_invoice') or {},
        }
    return None


# ─────────────────────────────────────────────
#  Orquestador principal — fase 1: escaneo
# ─────────────────────────────────────────────

def escanear_facturas_gmail_para_panel(fecha_desde: str | None = None) -> dict:
    """
    Escanea Gmail, encola facturas en facturas_compra_pendientes.json
    y devuelve resultado estructurado para el panel (sin WhatsApp).

    Por defecto solo factura del año en curso (fecha DIAN).
    Relee adjuntos aunque ya se hayan descargado antes (si no están en cola
    ni en historial), para no perder facturas del año al podar la cola.
    """
    anio_actual = datetime.now().year
    if not fecha_desde:
        fecha_desde = f"{anio_actual}/01/01"
    from app.tools.sincronizar_facturas_de_compra_siigo import (
        GmailAuthError,
        get_gmail_service,
        leer_correos_facturas_periodo,
        descargar_y_extraer_zip,
        extraer_xml_de_zip_local,
        extraer_datos_xml_dian,
    )

    resultado = {
        "ok": False,
        "correos_revisados": 0,
        "encoladas": [],
        "ya_en_cola": [],
        "ya_en_historial": [],
        "omitidas": [],
        "errores": [],
        "mensaje": "",
        "anio": anio_actual,
    }

    try:
        # Incluir ZIPs ya bajados: el manifiesto "descargadas" no implica "encoladas".
        correos = leer_correos_facturas_periodo(
            fecha_desde=fecha_desde,
            solo_no_descargados=False,
        )
    except GmailAuthError as e:
        resultado["mensaje"] = str(e)
        return resultado

    resultado["correos_revisados"] = len(correos)
    if not correos:
        resultado["ok"] = True
        resultado["mensaje"] = (
            f"No hay correos con ZIP en FACTURAS-MCKG desde {fecha_desde} "
            f"(año {anio_actual})."
        )
        return resultado

    try:
        service = get_gmail_service()
    except GmailAuthError as e:
        resultado["mensaje"] = str(e)
        return resultado

    compras_siigo = obtener_compras_siigo_para_dedupe()
    state = _cargar_pendientes()
    numeros_en_cola = {
        (v.get("numero_factura") or "").upper()
        for v in state.get("pendientes", {}).values()
    }
    sincronizar_historial_desde_importaciones()
    numeros_historial = {
        str(r.get("numero_factura") or "").strip().upper()
        for r in (_cargar_historial().get("historial") or [])
        if r.get("numero_factura")
    }

    for correo in correos:
        asunto = correo.get("asunto", "Sin asunto")
        for adjunto in correo.get("adjuntos_zip") or []:
            try:
                xml_content = None
                zip_local = adjunto.get("zip_path")
                if zip_local and os.path.isfile(zip_local):
                    xml_content, _pdf, _pdf_name = extraer_xml_de_zip_local(zip_local)
                if not xml_content:
                    xml_content, _pdf, _pdf_name = descargar_y_extraer_zip(
                        service, correo["id"], adjunto["id"], adjunto["filename"]
                    )
                if not xml_content:
                    resultado["errores"].append({
                        "asunto": asunto,
                        "archivo": adjunto.get("filename", ""),
                        "motivo": "ZIP sin XML DIAN válido",
                    })
                    continue

                datos = extraer_datos_xml_dian(xml_content)
                if not datos:
                    resultado["errores"].append({
                        "asunto": asunto,
                        "archivo": adjunto.get("filename", ""),
                        "motivo": "No se pudo parsear el XML DIAN",
                    })
                    continue

                numero_factura = f"{datos['prefix']}{datos['number']}"
                proveedor = datos.get("proveedor", "")
                nit = datos.get("nit_proveedor") or datos.get("nit") or ""
                fecha_fac = (datos.get("fecha") or "")[:10]
                if not fecha_fac.startswith(str(anio_actual)):
                    resultado["omitidas"].append({
                        "numero_factura": numero_factura,
                        "proveedor": proveedor,
                        "motivo": f"Fuera de {anio_actual} (fecha {fecha_fac or 'desconocida'})",
                    })
                    continue
                n_items = len(datos.get("items", []))
                total = round(sum(
                    item.get("subtotal", 0) + sum(
                        imp.get("valor", 0) for imp in item.get("impuestos", [])
                    )
                    for item in datos.get("items", [])
                ), 2)

                num_u = numero_factura.upper()
                if num_u in numeros_en_cola:
                    resultado["ya_en_cola"].append({
                        "numero_factura": numero_factura,
                        "proveedor": proveedor,
                    })
                    continue

                if num_u in numeros_historial:
                    resultado["ya_en_historial"].append({
                        "numero_factura": numero_factura,
                        "proveedor": proveedor,
                        "fecha": fecha_fac,
                    })
                    continue

                compra_registrada = buscar_compra_siigo_registrada(datos, compras_siigo)
                if compra_registrada:
                    doc = compra_registrada.get("name") or compra_registrada.get("id") or "SIIGO"
                    resultado["omitidas"].append({
                        "numero_factura": numero_factura,
                        "proveedor": proveedor,
                        "motivo": f"Ya registrada en SIIGO ({doc})",
                    })
                    continue

                es_nuevo = not es_proveedor_especial(nit, proveedor)
                sufijo = _encolar_factura(numero_factura, datos, xml_content, es_nuevo)
                numeros_en_cola.add(num_u)
                resultado["encoladas"].append({
                    "sufijo": sufijo,
                    "numero_factura": numero_factura,
                    "proveedor": proveedor,
                    "nit": nit,
                    "items_count": n_items,
                    "total": total,
                    "fecha": fecha_fac,
                    "es_nuevo_proveedor": es_nuevo,
                    "estado": "esperando_clasificacion" if es_nuevo else "esperando_confirmacion",
                })
            except Exception as e:
                resultado["errores"].append({
                    "asunto": asunto,
                    "archivo": adjunto.get("filename", ""),
                    "motivo": str(e)[:200],
                })

    n_enc = len(resultado["encoladas"])
    n_omit = len(resultado["omitidas"])
    n_hist = len(resultado["ya_en_historial"])
    n_cola = len(resultado["ya_en_cola"])
    if n_enc:
        resultado["ok"] = True
        resultado["mensaje"] = (
            f"{n_enc} factura(s) de {anio_actual} lista(s) para revisión. "
            "Revísalas una a una antes de confirmar."
        )
    elif resultado["errores"] and not (n_omit or n_hist or n_cola):
        resultado["mensaje"] = "Se encontraron correos pero hubo errores al leer los XML."
    elif n_omit or n_hist or n_cola:
        resultado["ok"] = True
        partes = []
        if n_hist:
            partes.append(f"{n_hist} ya en historial")
        if n_cola:
            partes.append(f"{n_cola} ya en cola")
        if n_omit:
            partes.append(f"{n_omit} omitida(s)")
        resultado["mensaje"] = (
            f"No hay facturas nuevas de {anio_actual} para cargar "
            f"({'; '.join(partes)})."
        )
    else:
        resultado["ok"] = True
        resultado["mensaje"] = f"No se encontraron facturas de {anio_actual} en Gmail."

    return resultado


def procesar_facturas_para_importar_productos(dias: int = 30) -> str:
    """
    Fase 1: Lee facturas del correo y las encola para aprobación manual por WhatsApp.
    Para cada factura detectada envía un mensaje al grupo preguntando si es
    materia prima (inventariar) o gasto/consumible — sin procesar nada aún.

    La fase 2 (procesamiento real) ocurre cuando el operador responde con:
      inv ok <código>          → proveedor conocido, procesar
      inv skip <código>        → omitir esta factura
      inv inventario <código>  → proveedor nuevo, tratar como materia prima
      inv gasto <código>       → proveedor nuevo, tratar como gasto/consumible
    """
    print(f"\n🚀 [IMPORTACIÓN] Escaneando facturas de proveedor en Gmail...")
    scan = escanear_facturas_gmail_para_panel()
    if not scan.get("ok") and scan.get("mensaje"):
        return f"❌ {scan['mensaje']}"

    encoladas = scan.get("encoladas") or []
    for f in encoladas:
        print(f"  📥 Encolada: {f['numero_factura']} ({f['proveedor']}) — #{f['sufijo']}")
    for o in scan.get("omitidas") or []:
        print(f"  ⏭️ Omitida: {o['numero_factura']} — {o.get('motivo', '')}")
    for e in scan.get("errores") or []:
        print(f"  ⚠️ Error: {e.get('asunto', '')} — {e.get('motivo', '')}")

    if not encoladas:
        return scan.get("mensaje") or "No se encontraron facturas nuevas."

    lineas = [f"⏳ {f['numero_factura']} ({f['proveedor']}) — #{f['sufijo']}" for f in encoladas]
    return (
        f"✅ {len(encoladas)} factura(s) en cola del panel. Ábrelas en Facturas de Compra.\n"
        + "\n".join(lineas)
    )


# ─────────────────────────────────────────────
#  Orquestador principal — fase 2: respuesta
# ─────────────────────────────────────────────

def procesar_respuesta_factura_compra(comando: str, sufijo: str, origen: str = 'whatsapp') -> str:
    """
    Fase 2: Procesa la respuesta del operador a una factura pendiente.

    Comandos válidos (llegan desde el grupo de contabilidad):
      ok          → proveedor ya registrado, procesar factura
      skip        → omitir esta factura (no procesar)
      inventario  → proveedor nuevo, clasificar como materia prima + procesar
                    (se agrega automáticamente a proveedores_especiales.json)
      gasto       → proveedor nuevo, clasificar como gasto/consumible (no procesar)
    """
    key, entrada = _buscar_pendiente(sufijo)
    if not entrada:
        return f"⚠️ No encontré factura pendiente con código *{sufijo}*.\nUsa *inv lista* para ver las pendientes."

    numero_factura = entrada['numero_factura']
    proveedor      = entrada['proveedor']
    nit            = entrada['nit']
    datos          = json.loads(entrada['datos_json'])
    cmd = comando.strip().lower()

    # ── Omitir / Gasto ───────────────────────────────────────────
    if cmd == 'skip':
        registrar_factura_historial(
            entrada,
            'omitida',
            sufijo=key,
            origen=origen,
            datos=datos,
            mensaje=f'Factura {numero_factura} omitida',
        )
        _quitar_pendiente(key)
        threading.Timer(4, _notificar_siguiente_factura_pendiente).start()
        return f"⏭️ Factura *{numero_factura}* omitida. No se registró nada en SIIGO."

    if cmd == 'gasto':
        total = entrada.get('total', 0)
        n_items = entrada.get('items_count', 0)

        from app.services.siigo import crear_factura_compra_siigo
        # Modelo basado en FC-1-42: document 5809 (FC), cost_center 263 (VENTAS), payment 1338
        payload_siigo = {
            "document": {"id": 5809},
            "date": datos.get("fecha", datetime.now().strftime("%Y-%m-%d")),
            "cost_center": 263,
            "supplier": {"identification": datos.get("nit", "999999999"), "branch_office": 0},
        "provider_invoice": {
            "prefix": datos.get("prefix") or "FV",
            "number": datos.get("number", "0")
        },
            "items": [{
                "type": "Account",
                "code": "11051001",
                "description": f"{proveedor} — {numero_factura}"[:100],
                "quantity": 1,
                "price": datos.get("total_neto", total),
                "taxes": []
            }],
            "payments": [{"id": 1338, "value": datos.get("total_neto", total)}],
            "observations": f"Gasto/consumible — {numero_factura} — {proveedor}"
        }

        resultado = crear_factura_compra_siigo(payload_siigo)

        if resultado.get("status") == "success":
            siigo_id = resultado.get("data", {}).get("id", "—")
            registrar_factura_historial(
                entrada,
                'gasto',
                sufijo=key,
                origen=origen,
                datos=datos,
                estado='ok',
                siigo_id=str(siigo_id),
                mensaje=f'Gasto registrado en SIIGO ({siigo_id})',
            )
        else:
            error = resultado.get("message", str(resultado))
            registrar_factura_historial(
                entrada,
                'gasto',
                sufijo=key,
                origen=origen,
                datos=datos,
                estado='error',
                mensaje=str(error)[:500],
            )

        _quitar_pendiente(key)
        threading.Timer(4, _notificar_siguiente_factura_pendiente).start()

        if resultado.get("status") == "success":
            siigo_id = resultado.get("data", {}).get("id", "—")
            return (
                f"✅ *Factura {numero_factura} registrada en SIIGO*\n"
                f"🏢 Proveedor: {proveedor}\n"
                f"📦 {n_items} ítem(s)  |  💰 Total: ${total:,.0f} COP\n"
                f"🆔 ID SIIGO: {siigo_id}"
            )
        else:
            error = resultado.get("message", str(resultado))
            return (
                f"❌ *Error al registrar {numero_factura} en SIIGO*\n"
                f"🏢 Proveedor: {proveedor}\n"
                f"⚠️ Error: {error[:200]}\n\n"
                f"Registra manualmente: SIIGO → Compras → Nueva compra o gasto"
            )

    # ── Inventario (proveedor nuevo) ─────────────────────────────
    if cmd == 'inventario':
        # Agregar automáticamente a la lista de proveedores especiales
        data_prov = cargar_proveedores_especiales()
        nit_limpio = re.sub(r'\D', '', nit or '')
        ya_existe = any(
            re.sub(r'\D', '', p.get('nit', '')) == nit_limpio
            for p in data_prov.get('proveedores', [])
            if nit_limpio
        )
        if not ya_existe:
            data_prov['proveedores'].append({
                'nit':    nit,
                'nombre': proveedor,
                'activo': True,
                'nota':   f'Agregado automáticamente el {datetime.now().strftime("%Y-%m-%d")} vía WhatsApp',
            })
            with open(_RUTA_PROVEEDORES, 'w', encoding='utf-8') as f:
                json.dump(data_prov, f, indent=2, ensure_ascii=False)
            print(f"📋 [IMPORTACIÓN] Proveedor agregado a lista especial: {proveedor} ({nit})")

    # ── Procesar (ok o inventario) ───────────────────────────────
    if cmd in ('ok', 'inventario'):
        xml_content = base64.b64decode(entrada['xml_b64']).decode('utf-8')

        _quitar_pendiente(key)

        nota_prov = (
            f"\n✅ *{proveedor}* agregado a la lista de proveedores especiales."
            if cmd == 'inventario' and not ya_existe
            else ""
        ) if cmd == 'inventario' else ""

        enviar_whatsapp_reporte(
            f"⚙️ Procesando factura *{numero_factura}* ({proveedor})…\n"
            f"Generando códigos McKenna, Excel y XML de compra SIIGO.{nota_prov}",
            numero_destino=GRUPO_COMPRAS,
        )

        arch = _ejecutar_procesamiento(numero_factura, datos, xml_content)
        if not arch:
            registrar_factura_historial(
                entrada,
                'inventario',
                sufijo=key,
                origen=origen,
                datos=datos,
                estado='error',
                mensaje='Sin ítems procesables',
            )
            threading.Timer(4, _notificar_siguiente_factura_pendiente).start()
            return f"⚠️ Factura *{numero_factura}*: no se encontraron ítems procesables."

        registrar_factura_historial(
            entrada,
            'inventario',
            sufijo=key,
            origen=origen,
            datos=datos,
            items_resumen=_resumen_items_historial(arch.get('productos')),
            nuevos=arch.get('nuevos', 0),
            en_siigo=arch.get('duplicados', 0),
            ruta_excel=arch.get('ruta'),
            ruta_xml=arch.get('ruta_xml'),
            mensaje=f"Excel + XML generados ({arch.get('nuevos', 0)} nuevos)",
        )
        # Notificar la siguiente factura en cola (si existe) con pausa
        threading.Timer(4, _notificar_siguiente_factura_pendiente).start()

        return (
            f"✅ *Factura {numero_factura} procesada*\n"
            f"🏢 {proveedor}\n"
            f"📦 Nuevos: {arch['nuevos']} | Duplicados: {arch['duplicados']}\n"
            f"📎 Excel + XML enviados al grupo."
        )

    return f"⚠️ Comando no reconocido: *{comando}*. Usa: inv ok / inv skip / inv inventario / inv gasto"


def listar_facturas_pendientes() -> str:
    """Retorna un resumen de las facturas en cola de aprobación."""
    state = _cargar_pendientes()
    pendientes = state.get('pendientes', {})
    if not pendientes:
        return "✅ No hay facturas de compra pendientes de clasificación."
    lineas = ["📋 *Facturas pendientes de clasificación:*\n"]
    for sufijo, e in pendientes.items():
        estado = "❓ nuevo proveedor" if e.get('es_nuevo_proveedor') else "✅ proveedor conocido"
        lineas.append(
            f"  • *{sufijo}* — {e['numero_factura']} | {e['proveedor']} | "
            f"{e['items_count']} ítems | ${e['total']:,.0f} COP | {estado}"
        )
    return "\n".join(lineas)


def _construir_resumen_whatsapp(arch: dict) -> str:
    xml_nombre  = os.path.basename(arch.get('ruta_xml', '')) or '—'
    excel_nombre = os.path.basename(arch.get('ruta', '')) or '—'
    return (
        f"📦 *FLUJO DE COMPRA — Registro de Inventario SIIGO*\n"
        f"_(Independiente del flujo de facturación de venta a clientes)_\n\n"
        f"🔢 *Factura proveedor:* {arch['numero_factura']}\n"
        f"🏢 *Proveedor:* {arch['proveedor']}\n\n"
        f"✅ *Productos nuevos:* {arch['nuevos']}\n"
        f"📦 *Ya en SIIGO (suman inventario):* {arch['duplicados']}\n\n"
        f"📎 *Archivos generados:*\n"
        f"   • Excel: `{excel_nombre}`\n"
        f"   • XML:   `{xml_nombre}`\n\n"
        f"📋 *Protocolo de carga en SIIGO:*\n"
        f"\n"
        f"   *— Paso A: Registrar los productos nuevos (Excel) —*\n"
        f"   1. SIIGO → Inventario → Productos\n"
        f"   2. Clic en ▶ *Importación*\n"
        f"   3. Selecciona el Excel adjunto\n"
        f"   4. Verifica la vista previa y confirma\n"
        f"   _(Omitir si no hay productos nuevos)_\n\n"
        f"   *— Paso B: Registrar la compra (XML) —*\n"
        f"   1. SIIGO → Compras\n"
        f"   2. Clic en ▶ *Crear compra o gasto desde un XML o ZIP*\n"
        f"   3. Carga el archivo XML adjunto\n"
        f"   4. Verifica que el total coincide con la factura del proveedor\n"
        f"   5. Asienta el documento — los ítems en SIIGO suman inventario\n\n"
        f"ℹ️ *Nota:* Los productos que ya existen en SIIGO solo van en el XML de compra."
    )


if __name__ == "__main__":
    resultado = procesar_facturas_para_importar_productos()
    print(f"\n{resultado}")
